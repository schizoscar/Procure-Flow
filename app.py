# app.py
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file, abort, g
from models import db, User, Supplier, Category, CategoryItem, Task, PRItem, TaskSupplier, SupplierQuote, FileAsset, EmailLog
import re
from datetime import datetime
import json
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
import requests
import io
from dotenv import load_dotenv
import imaplib
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime
import threading
import time
from itsdangerous import URLSafeSerializer, BadSignature
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Email, To, Content
import certifi
from urllib.parse import urljoin
import logging
from werkzeug.exceptions import HTTPException
import math
from decimal import Decimal, InvalidOperation
from sqlalchemy import func, text, desc, and_, or_, cast, String, Integer, Float, DateTime, Boolean
from sqlalchemy.exc import SQLAlchemyError

# Load .env file for local development
load_dotenv()

# Check if running on Render
IS_RENDER = os.getenv('RENDER', '0') == '1'

app = Flask(__name__)
secret = os.getenv("APP_SECRET_KEY")
if not secret:
    raise RuntimeError("APP_SECRET_KEY is not set")
app.secret_key = secret

# File upload configuration
UPLOADS_DIR = os.path.join('uploads', 'certificates')
ALLOWED_EXTENSIONS = {'pdf'}  # PDF only
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
app.config["PUBLIC_BASE_URL"] = PUBLIC_BASE_URL

# Ensure uploads directory exists
os.makedirs(UPLOADS_DIR, exist_ok=True)

# Basic logging setup (works for dev + prod)
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s"
)

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file_obj, task_id, supplier_id, pr_item_id):
    """
    Store uploaded file into DB and return file_assets.id
    """
    filename = secure_filename(file_obj.filename or "document.pdf")
    mime_type = file_obj.mimetype or "application/pdf"
    data = file_obj.read()  # bytes

    file_asset = FileAsset(
        task_id=task_id,
        supplier_id=supplier_id,
        pr_item_id=pr_item_id,
        filename=filename,
        mime_type=mime_type,
        data=data
    )
    
    db.session.add(file_asset)
    db.session.flush()  # Get the ID without committing
    return file_asset.id

def public_url(endpoint: str, **values) -> str:
    """
    Build a fully-qualified URL using PUBLIC_BASE_URL (preferred),
    otherwise fall back to Flask's _external=True behavior.
    """
    base = app.config.get("PUBLIC_BASE_URL")
    if base:
        # build path only, then prefix with base
        path = url_for(endpoint, _external=False, **values)
        return f"{base}{path}"
    # fallback (will be localhost if you run locally)
    return url_for(endpoint, _external=True, **values)

def public_url_for(endpoint: str, **values) -> str:
    """
    Builds a public absolute URL.
    - If PUBLIC_BASE_URL is set (recommended), it forces that domain.
    - Otherwise falls back to Flask's _external=True (local dev).
    """
    # Always generate a path first
    path = url_for(endpoint, **values)  # e.g. "/file/12"
    if PUBLIC_BASE_URL:
        return urljoin(PUBLIC_BASE_URL + "/", path.lstrip("/"))
    return url_for(endpoint, _external=True, **values)

@app.route('/file/<int:file_id>')
def serve_file(file_id):
    """Serve a file from the database."""
    file_asset = db.session.get(FileAsset, file_id)
    
    if not file_asset:
        return "File not found", 404
        
    return send_file(
        io.BytesIO(file_asset.data),
        mimetype=file_asset.mime_type,
        download_name=file_asset.filename,
        as_attachment=False
    )

def get_quote_serializer():
    return URLSafeSerializer(app.secret_key, salt="supplier-quote")

def get_reset_serializer():
    return URLSafeSerializer(app.secret_key, salt="password-reset")

def make_reset_token(user_id: int, email: str) -> str:
    return get_reset_serializer().dumps({"user_id": user_id, "email": email})

def verify_reset_token(token: str):
    return get_reset_serializer().loads(token)  # we'll enforce expiry via our own timestamp check if desired

def generate_temp_password(length: int = 10) -> str:
    # simple temp password: letters+digits (no symbols to avoid email copy issues)
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(alphabet[uuid.uuid4().int % len(alphabet)] for _ in range(length))

# Database configuration
database_url = os.getenv('DATABASE_URL')
if not database_url:
    database_url = 'sqlite:///database/procure_flow.db'
elif database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Create tables on startup
with app.app_context():
    db.create_all()
    print("Database tables created/verified")

# Email Configuration
EMAIL_CONFIG = {
    'smtp_server': os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
    'smtp_port': int(os.getenv('SMTP_PORT', '587')),
    'sender_email': os.getenv('SMTP_SENDER'),
    'sender_password': os.getenv('SMTP_PASSWORD')
}
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
SENDGRID_SENDER = os.environ.get("SENDGRID_SENDER", "")

# IMAP configuration (for inbox polling)
IMAP_SERVER = os.getenv('IMAP_SERVER', 'imap.gmail.com')
IMAP_PORT = int(os.getenv('IMAP_PORT', '993'))
IMAP_USERNAME = os.getenv('IMAP_USERNAME', EMAIL_CONFIG['sender_email'])
IMAP_PASSWORD = os.getenv('IMAP_PASSWORD', EMAIL_CONFIG['sender_password'])

ENABLE_DEBUG_ROUTES = os.getenv("ENABLE_DEBUG_ROUTES", "0") == "1"

# ==================================== START OF MIGRATION ====================================
'''
def check_database_status():
    """Check if database needs migration."""
    with app.app_context():
        try:
            from sqlalchemy import func
            
            # Just check suppliers - you know you should have 543
            supplier_count = db.session.query(func.count(Supplier.id)).scalar()
            
            if supplier_count < 10:  # Very low threshold
                print(f"⚠️ Only {supplier_count} suppliers found (expected ~543)")
                print("Visit /admin/migrate to run migration manually")
                return False
            else:
                print(f"✅ Database OK: {supplier_count} suppliers")
                return True
                
        except Exception as e:
            print(f"⚠️ Database check failed: {e}")
            return False

# Run check on startup
import threading
thread = threading.Thread(target=check_database_status, daemon=True)
thread.start()
'''
# Add admin migration routes
@app.route('/admin/migrate')
def admin_migrate_page():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    from sqlalchemy import func
    supplier_count = db.session.query(func.count(Supplier.id)).scalar()
    
    return f'''
    <h2>Database Migration</h2>
    <p>Current suppliers: {supplier_count}</p>
    <p>Expected suppliers: 543</p>
    <form method="POST" action="/admin/migrate/run">
        <button type="submit">Run Migration</button>
    </form>
    '''

@app.route('/admin/migrate/run', methods=['POST'])
def run_migration_admin():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    try:
        if os.path.exists('database/procure_flow.db'):
            from migrate_sqlite_to_postgres import migrate_all_data
            migrate_all_data()
            return "✅ Migration completed! <a href='/'>Go to dashboard</a>"
        else:
            return "❌ SQLite file not found"
    except Exception as e:
        return f"❌ Migration failed: {str(e)}"
# ==================================== END OF MIGRATION ==============================

@app.template_filter('format_date')
def format_date_filter(value, format='%Y-%m-%d'):
    """Format a datetime object for display."""
    if value is None:
        return 'N/A'
    if isinstance(value, str):
        # Handle string dates (fallback for compatibility)
        return value[:10] if len(value) >= 10 else value
    try:
        return value.strftime(format)
    except AttributeError:
        return str(value)

# Validation functions
def validate_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_phone(phone):
    # Malaysian phone number validation
    pattern = r'^(\+?6?01)[0-46-9]-*[0-9]{7,8}$'
    return re.match(pattern, phone.replace(' ', '').replace('-', '')) is not None

def validate_password(password):
    # At least 5 letters and 1 number
    if len(password) < 6:
        return False
    if not any(char.isdigit() for char in password):
        return False
    if not any(char.isalpha() for char in password):
        return False
    return True

def generate_email_content(pr_items, task_name):
    # Decide header label based on the category used in this PR
    first_cat = ""
    for it in (pr_items or []):
        d = dict(it) if it is not None else {}
        c = (d.get("item_category") or "").strip()
        if c:
            first_cat = c
            break

    DIM_CATS = {"Steel Plates", "Stainless Steel", "Angle Bar", "Rebar", "Bolts, Fasteners"}
    is_other = first_cat not in DIM_CATS

    dim_header = "Dimensions" if not is_other else "Packing"
    qty_header = "Qty (UOM)" if is_other else "Qty"

    items_html = f"""
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; border: 1px solid #ddd; font-family: Arial, sans-serif;">
        <thead style="background-color: #f2f2f2;">
            <tr>
                <th style="text-align: center; width: 5%;">No.</th>
                <th style="text-align: left;">Description</th>
                <th style="text-align: left;">{dim_header}</th>
                <th style="text-align: left;">Brand / Specification</th>
                <th style="text-align: center; width: 10%;">{qty_header}</th>
                <th style="text-align: left;">Remark</th>
            </tr>
        </thead>
        <tbody>
    """

    for idx, item in enumerate(pr_items, 1):
        item = dict(item) if item is not None else {}
        category = (item.get('item_category') or '').strip()

        # Build Specification (Dimensions/Packing content)
        spec = ''
        if category in ['Steel Plates', 'Stainless Steel']:
            w = item.get('width') or ''
            l = item.get('length') or ''
            thk = item.get('thickness') or ''
            if w or l or thk:
                spec = f"{w} mm (W) x {l} mm (L) x {thk} mm (Thk)"

        elif category == 'Angle Bar':
            a = item.get('dim_a') or ''
            b = item.get('dim_b') or ''
            l = item.get('length') or ''
            thk = item.get('thickness') or ''
            if a or b or l or thk:
                spec = f"{a} mm (A) x {b} mm (B) x {l} mm (L) x {thk} mm (Thk)"

        elif category in ['Rebar', 'Bolts, Fasteners']:
            d = item.get('diameter') or ''
            l = item.get('length') or ''
            if d or l:
                spec = f"{d} mm (D) x {l} mm (L)"

        else:
            # "Others" -> Packing should use ONLY uom_qty now
            uom_qty = (item.get('uom_qty') or '').strip() if isinstance(item.get('uom_qty'), str) else (item.get('uom_qty') or '')
            if uom_qty:
                spec = f"{uom_qty}"
            else:
                spec = ""

        original_spec = item.get('specification') or ''
        if spec and original_spec:
            final_spec = f"{spec}<br><small>{original_spec}</small>"
        elif spec:
            final_spec = spec
        else:
            final_spec = original_spec or 'N/A'

        # Qty display: for Others show "quantity + uom", else just quantity
        qty_val = item.get('quantity') or ''
        if category not in DIM_CATS:
            uom = (item.get('uom') or '').strip()
            qty_display = f"{qty_val} {uom}".strip() if (qty_val or uom) else ''
        else:
            qty_display = f"{qty_val}"

        items_html += f"""
            <tr>
                <td style="text-align: center;">{idx}</td>
                <td>{item.get('item_name') or ''}</td>
                <td>{final_spec}</td>
                <td>{item.get('brand') or ''}</td>
                <td style="text-align: center;">{qty_display}</td>
                <td>{item.get('our_remarks') or ''}</td>
            </tr>
        """

    items_html += "</tbody></table>"

    return f"""
    <html>
    <body>
        <h2>Procurement Inquiry</h2>
        <p>Dear {{supplier_name}}{{contact_person}},</p>

        <p>We are inquiring about the following items for procurement:</p>
        
        {items_html}

        <p>Please provide us with your quotation including:</p>
        <ul>
            <li>Payment Terms (Days)</li>
            <li>Unit Price (RM)</li>
            <li>Delivery Lead Timeline (Days)</li>
            <li>Stock Availability</li>
            <li>Warranty (If Applicable)</li>
            <li>Mill Certificate / Certificate of Analysis (COA)</li>
        </ul>
        
        <p>Please fill in the quotation in the link below:</p>
        <p>Supplier form:</p>
        <p>↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓</p>
        {{quote_form_link}}
        <p>↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑</p>
        
        <p>We look forward to your prompt response.</p>
        
        <p>Best regards,<br>
        Procurement Department</p>
    </body>
    </html>
    """

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    """Main dashboard showing recent tasks and stats."""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # Subquery for item counts
    item_count_subq = db.session.query(
        PRItem.task_id,
        func.count(PRItem.id).label('item_count')
    ).group_by(PRItem.task_id).subquery()

    # Get recent tasks with joins
    recent_tasks_query = db.session.query(
        Task,
        User.username.label('created_by'),
        func.coalesce(item_count_subq.c.item_count, 0).label('item_count')  # Use func.coalesce
    ).join(
        User, Task.user_id == User.id
    ).outerjoin(
        item_count_subq, Task.id == item_count_subq.c.task_id
    ).order_by(
        Task.created_at.desc()
    ).limit(10).all()

    # Convert to list of dictionaries
    recent_tasks = []
    for task, created_by, item_count in recent_tasks_query:
        recent_tasks.append({
            'id': task.id,
            'task_name': task.task_name,
            'user_id': task.user_id,
            'status': task.status,
            'created_at': task.created_at,
            'created_by': created_by,
            'item_count': item_count
        })

    # Get stats using SQLAlchemy queries
    stats = {
        'total_tasks': db.session.query(Task).count(),
        'active_tasks': db.session.query(Task).filter(
            ~Task.status.in_(['completed', 'cancelled'])
        ).count(),
        'total_suppliers': db.session.query(Supplier).filter_by(is_active=True).count()
    }

    return render_template('dashboard.html', recent_tasks=recent_tasks, stats=stats)

@app.route('/purchase-requisitions')
def purchase_requisitions():
    """Show saved Purchase Requisitions that haven't been sent to suppliers yet."""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    prs = db.session.query(
        Task,
        User.username.label('created_by'),
        func.count(PRItem.id).label('item_count')  # Use func.count
    ).join(
        User, Task.user_id == User.id
    ).outerjoin(
        PRItem, Task.id == PRItem.task_id
    ).filter(
        Task.status == 'purchase_requisition'
    ).group_by(
        Task.id, User.username
    ).order_by(
        Task.created_at.desc()
    ).all()

    prs_list = []
    for task, created_by, item_count in prs:
        pr_dict = {
            'id': task.id,
            'task_name': task.task_name,
            'user_id': task.user_id,
            'status': task.status,
            'created_at': task.created_at,
            'created_by': created_by,
            'item_count': item_count
        }
        prs_list.append(pr_dict)

    return render_template('purchase_requisitions.html', prs=prs_list)

@app.route('/uploads/certificates/<path:filepath>')
def download_certificate(filepath):
    """Serve uploaded certificate files."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    try:
        file_path = os.path.join(UPLOADS_DIR, filepath)
        # Prevent directory traversal
        if not os.path.abspath(file_path).startswith(os.path.abspath(UPLOADS_DIR)):
            return "Access denied", 403
        
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return "File not found", 404
    except Exception as e:
        print(f"Error serving certificate: {e}")
        return "Error accessing file", 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = db.session.query(User).filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('login.html')

@app.route('/create-user', methods=['GET', 'POST'])
def create_user():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form['username'].strip()
        email_addr = request.form['email'].strip().lower()
        password = (request.form.get('password') or "").strip()
        role = request.form['role']

        # Validation (keep your existing validation)
        if not validate_email(email_addr):
            flash('Invalid email format', 'error')
            return render_template('create_user.html')

        # If password empty, generate temp password
        is_temp_password = False
        if not password:
            password = generate_temp_password(10)
            is_temp_password = True

        if not validate_password(password):
            flash('Password must contain at least 5 letters and 1 number', 'error')
            return render_template('create_user.html')

        password_hash = generate_password_hash(password)

        try:
            # Check if user exists
            existing_user = db.session.query(User).filter(
                (User.username == username) | (User.email == email_addr)
            ).first()
            
            if existing_user:
                flash('Username or email already exists', 'error')
                return render_template('create_user.html')
            
            # Create new user
            new_user = User(
                username=username,
                email=email_addr,
                password_hash=password_hash,
                role=role
            )
            
            db.session.add(new_user)
            db.session.flush()  # Get the ID
            new_user_id = new_user.id

            # Build reset link (keep your existing email sending logic)
            token = make_reset_token(new_user_id, email_addr)
            reset_link = public_url_for("reset_password", token=token)

            # Email login details + reset link
            subject = "Your Procure-Flow account details"
            html = f"""
            <html><body>
            <p>Hello {username},</p>
            <p>An account has been created for you in Procure-Flow.</p>
            <p><strong>Login details:</strong><br>
            Username: <code>{username}</code><br>
            Email: <code>{email_addr}</code><br>
            Password: <code>{password}</code></p>
            <p><strong>Reset your password now (recommended):</strong><br>
            <a href="{reset_link}">{reset_link}</a></p>
            </body></html>
            """

            email_ok = send_email_html(email_addr, subject, html, to_name=username)
            if email_ok:
                flash('User created successfully and email sent.', 'success')
            else:
                flash('User created, but failed to send email. Please verify email settings.', 'error')

            db.session.commit()
            return redirect(url_for('index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating user: {str(e)}', 'error')
            return render_template('create_user.html')

    return render_template('create_user.html')

@app.route('/user-list')
def user_list():
    """Display all users for admin management."""
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))

    users = db.session.query(User).order_by(User.created_at.desc()).all()
    
    return render_template('user_list.html', users=users)

@app.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
def edit_user(user_id):
    """Edit user details (admin only)."""
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))

    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('user_list'))

    if request.method == 'POST':
        new_username = request.form['username'].strip()
        new_email = request.form['email'].strip().lower()
        new_role = request.form['role']
        new_password = (request.form.get('password') or "").strip()
        confirm_password = (request.form.get('confirm_password') or "").strip()

        # Validation
        if not validate_email(new_email):
            flash('Invalid email format', 'error')
            return render_template('edit_user.html', user=user)

        # Check if password change requested
        if new_password or confirm_password:
            if new_password != confirm_password:
                flash('Passwords do not match', 'error')
                return render_template('edit_user.html', user=user)

            if not validate_password(new_password):
                flash('Password must contain at least 5 letters and 1 number', 'error')
                return render_template('edit_user.html', user=user)

            password_hash = generate_password_hash(new_password)
            user.password_hash = password_hash

        # Update user fields
        user.username = new_username
        user.email = new_email
        user.role = new_role

        try:
            db.session.commit()
            flash('User updated successfully!', 'success')
            return redirect(url_for('user_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'error')
            return render_template('edit_user.html', user=user)

    return render_template('edit_user.html', user=user)

@app.route('/user/<int:user_id>/delete', methods=['POST'])
def delete_user(user_id):
    """Delete a user (admin only)."""
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('index'))

    MAIN_USER_ID = 1
    if user_id == MAIN_USER_ID:
        flash('The main admin account cannot be deleted.', 'error')
        return redirect(url_for('user_list'))

    user = db.session.get(User, user_id)
    
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('user_list'))

    try:
        # Check if user has created any tasks
        task_count = db.session.query(Task).filter_by(user_id=user_id).count()
        if task_count > 0:
            flash(f'Cannot delete user "{user.username}" because they have created {task_count} task(s).', 'error')
            return redirect(url_for('user_list'))
        
        db.session.delete(user)
        db.session.commit()
        flash(f'User "{user.username}" has been deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Cannot delete this user: {str(e)}', 'error')
    
    return redirect(url_for('user_list'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email_addr = (request.form.get('email') or '').strip().lower()
        if not validate_email(email_addr):
            flash("Please enter a valid email address.", "error")
            return render_template("forgot_password.html")

        user = db.session.query(User).filter(
            func.lower(User.email) == email_addr  # Use func.lower
        ).first()
        
        # Always respond success to avoid account enumeration
        if user:
            token = make_reset_token(user.id, user.email)
            reset_link = public_url_for("reset_password", token=token)

            subject = "Reset your Procure-Flow password"
            html = f"""
            <html><body>
            <p>Hello {user.username},</p>
            <p>Click the link below to reset your password:</p>
            <p><a href="{reset_link}">{reset_link}</a></p>
            <p>If you did not request this, you can ignore this email.</p>
            </body></html>
            """
            send_email_html(user.email, subject, html, to_name=user.username)

        flash("If an account exists for that email, a reset link has been sent.", "success")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        data = get_reset_serializer().loads(token)
        user_id = data.get("user_id")
        email_addr = (data.get("email") or "").lower()
        if not user_id or not email_addr:
            return render_template("errors/400.html"), 400
    except BadSignature:
        return render_template("errors/400.html"), 400

    user = db.session.query(User).filter(
        User.id == user_id,
        func.lower(User.email) == email_addr
    ).first()
    
    if not user:
        return render_template("errors/404.html"), 404

    if request.method == 'POST':
        new_password = (request.form.get("password") or "").strip()
        confirm = (request.form.get("confirm_password") or "").strip()

        if new_password != confirm:
            flash("Passwords do not match.", "error")
            return render_template("reset_password.html", username=user.username)

        if not validate_password(new_password):
            flash("Password must contain at least 5 letters and 1 number.", "error")
            return render_template("reset_password.html", username=user.username)

        new_hash = generate_password_hash(new_password)
        user.password_hash = new_hash
        
        try:
            db.session.commit()
            flash("Password reset successful. Please log in.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error resetting password: {str(e)}", "error")

    return render_template("reset_password.html", username=user.username)

def parse_int_field(value, *, required=False):
    s = (value or "").strip()
    if s == "":
        if required:
            raise ValueError("Required")
        return None
    if not s.isdigit():
        raise ValueError("Must be a whole number")
    return int(s)

def parse_optional_int(value):
    # Convenience: optional int, returns None if blank
    return parse_int_field(value, required=False)

DEC_RE = re.compile(r"^\d+(\.\d{1,2})?$")  # up to 2 dp

def parse_decimal_field(val, field_name):
    if val is None or val == "":
        return None
    if not DEC_RE.fullmatch(val):
        raise ValueError(f"{field_name} must be a number with up to 2 decimals")
    try:
        return Decimal(val)
    except InvalidOperation:
        raise ValueError(f"{field_name} is not a valid number")

@app.route('/new-task', methods=['GET', 'POST'])
@app.route('/edit-task/<int:task_id>', methods=['GET', 'POST'])
def new_task(task_id=None):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    categories = db.session.query(Category).order_by(Category.name).all()
    
    # Fetch category items for autocomplete
    cat_items_data = db.session.query(
        Category.name.label('cat_name'),
        CategoryItem.name.label('item_name')
    ).join(CategoryItem).order_by(CategoryItem.name).all()
    
    category_items_map = {}
    for row in cat_items_data:
        cat = row.cat_name
        if cat not in category_items_map:
            category_items_map[cat] = []
        category_items_map[cat].append(row.item_name)
    
    if task_id:
        # Editing existing task - verify ownership
        task = db.session.get(Task, task_id)
        if not task:
            flash('Task not found', 'error')
            return redirect(url_for('task_list'))
        
        # Get existing items
        existing_items = db.session.query(PRItem).filter_by(task_id=task_id).all()
    else:
        task = None
        existing_items = []
    
    if request.method == 'POST':
        project_ref = request.form['project_reference']
        global_category = request.form['global_category']
        
        # Format task name
        task_name = f"{project_ref} - {global_category}"
        
        items = []
        
        # Process items from form
        item_index = 0
        while f'items[{item_index}][item_name]' in request.form:
            try:
                qty = parse_int_field(request.form.get(f'items[{item_index}][quantity]'), required=True)

                width     = parse_optional_int(request.form.get(f'items[{item_index}][width]'))
                length    = parse_optional_int(request.form.get(f'items[{item_index}][length]'))
                thickness = parse_optional_int(request.form.get(f'items[{item_index}][thickness]'))

                dim_a = parse_optional_int(request.form.get(f'items[{item_index}][dim_a]'))
                dim_b = parse_optional_int(request.form.get(f'items[{item_index}][dim_b]'))

                diameter = parse_optional_int(request.form.get(f'items[{item_index}][diameter]'))

                uom_qty_raw = (request.form.get(f'items[{item_index}][uom_qty]') or '').strip()
                uom_qty = uom_qty_raw if uom_qty_raw else None

                our_remarks_raw = (request.form.get(f'items[{item_index}][our_remarks]') or '').strip()
                our_remarks = our_remarks_raw if our_remarks_raw else None

            except ValueError as e:
                flash(f"Item #{item_index+1}: invalid number input ({str(e)})", "error")
                return render_template('pr_form.html',
                                     categories=categories,
                                     task=task,
                                     existing_items=existing_items,
                                     is_edit=bool(task_id),
                                     category_items_map=category_items_map)

            items.append({
                'item_category': global_category,
                'item_name': request.form.get(f'items[{item_index}][item_name]'),
                'brand': request.form.get(f'items[{item_index}][brand]') or None,
                'quantity': qty,   # <-- now an int
                'payment_terms': request.form.get(f'items[{item_index}][payment_terms]') or None,

                'width': width,
                'length': length,
                'thickness': thickness,

                'dim_a': dim_a,
                'dim_b': dim_b,

                'diameter': diameter,

                'uom_qty': uom_qty,
                'uom': request.form.get(f'items[{item_index}][uom]') or None,

                'our_remarks': our_remarks,
            })
            item_index += 1
        
        try:
            if task_id:
                # Update existing task
                task.task_name = task_name
                # Delete existing items and add new ones
                db.session.query(PRItem).filter_by(task_id=task_id).delete()
                task_id_to_use = task_id
                flash('Task updated successfully!', 'success')
            else:
                # Create new task
                new_task = Task(
                    task_name=task_name,
                    user_id=session['user_id'],
                    status='purchase_requisition'
                )
                db.session.add(new_task)
                db.session.flush()
                task_id_to_use = new_task.id
                task = new_task  # Update task variable for template if needed
                flash('Purchase Requisition saved successfully!', 'success')
            
            # Add PR items with all dimension fields
            for item_data in items:
                pr_item = PRItem(
                    task_id=task_id_to_use,
                    item_category=item_data['item_category'],
                    item_name=item_data['item_name'],
                    brand=item_data['brand'],
                    quantity=item_data['quantity'],
                    payment_terms=item_data['payment_terms'],
                    width=item_data.get('width'),
                    length=item_data.get('length'),
                    thickness=item_data.get('thickness'),
                    dim_a=item_data.get('dim_a'),
                    dim_b=item_data.get('dim_b'),
                    diameter=item_data.get('diameter'),
                    uom_qty=item_data.get('uom_qty'),
                    uom=item_data.get('uom'),
                    our_remarks=item_data.get('our_remarks')
                )
                db.session.add(pr_item)
            
            db.session.commit()
            
            # Redirect to Purchase Requisitions list after save
            return redirect(url_for('purchase_requisitions'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving task: {str(e)}', 'error')
    
    return render_template('pr_form.html', 
                         categories=categories, 
                         task=task, 
                         existing_items=existing_items,
                         is_edit=bool(task_id),
                         category_items_map=category_items_map)

@app.route('/task/<int:task_id>/edit')
def edit_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    next_url = request.args.get("next")

    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))

    # Get the status while session is still available
    task_status = task.status

    def go(endpoint):
        return redirect(url_for(endpoint, task_id=task_id, next=next_url) if next_url else url_for(endpoint, task_id=task_id))

    if task_status == 'purchase_requisition':
        return go('new_task')
    elif task_status == 'select_suppliers':
        return go('supplier_selection')
    elif task_status == 'generate_email':
        return go('email_preview')
    elif task_status == 'confirm_email':
        return go('email_confirmation')
    else:
        return redirect(next_url) if next_url else redirect(url_for('task_list'))

@app.route('/task/<int:task_id>/email-preview', methods=['GET', 'POST'])
def email_preview(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verify task ownership
    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))
    
    # Get selected suppliers with their assigned items
    selected_suppliers = db.session.query(
        Supplier,
        TaskSupplier.assigned_items
    ).join(
        TaskSupplier, Supplier.id == TaskSupplier.supplier_id
    ).filter(
        TaskSupplier.task_id == task_id,
        TaskSupplier.is_selected == True
    ).all()
    
    pr_items = db.session.query(PRItem).filter_by(task_id=task_id).all()
    
    subject_key = f"email_subject_{task_id}"
    content_key = f"email_content_{task_id}"
    email_subject = session.get(subject_key, f"Procurement Inquiry - {task.task_name}")
    
    if request.method == 'POST':
        action = request.form.get('action')
        email_content = request.form.get('email_content', '')
        email_subject = request.form.get('email_subject', email_subject)
        
        if action == 'update_preview':
            session[content_key] = email_content
            session[subject_key] = email_subject
            flash('Preview updated!', 'success')
            
        elif action == 'send_emails':
            session['final_email_content'] = email_content
            session['final_email_subject'] = email_subject
            return redirect(url_for('email_confirmation', task_id=task_id))
    
    # Group suppliers by email template
    email_templates = {}
    for supplier, assigned_items in selected_suppliers:
        assigned_item_ids = None
        if assigned_items:
            try:
                assigned_item_ids = json.loads(assigned_items)
            except:
                assigned_item_ids = None
        
        if assigned_item_ids:
            key = tuple(sorted(assigned_item_ids))
        else:
            key = 'all'
        
        if key not in email_templates:
            email_content = session.get(content_key, '')
            if not email_content:
                # Filter items for this template
                template_items = [item for item in pr_items if not assigned_item_ids or item.id in assigned_item_ids]
                email_content = generate_email_content(template_items, task.task_name)
            
            email_templates[key] = {
                'suppliers': [],
                'items': assigned_item_ids if assigned_item_ids else [item.id for item in pr_items],
                'email_content': email_content
            }
        
        email_templates[key]['suppliers'].append({
            'id': supplier.id,
            'name': supplier.name,
            'email': supplier.email,
            'contact_name': supplier.contact_name,
            'assigned_items': assigned_items
        })
    
    return render_template('email_preview.html',
                         task=task,
                         email_templates=email_templates,
                         pr_items=pr_items,
                         email_subject=email_subject)

@app.route('/task/<int:task_id>/confirm-email', methods=['GET', 'POST'])
def email_confirmation(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))
    
    # Get selected suppliers
    selected_suppliers = db.session.query(
        Supplier,
        TaskSupplier.assigned_items
    ).join(
        TaskSupplier, Supplier.id == TaskSupplier.supplier_id
    ).filter(
        TaskSupplier.task_id == task_id,
        TaskSupplier.is_selected == True
    ).all()
    
    pr_items = db.session.query(PRItem).filter_by(task_id=task_id).all()

    # Mark task as in email generation stage
    if task.status != 'confirm_email':
        task.status = 'generate_email'
        db.session.commit()
    
    if request.method == 'POST':
        final_email_content = session.get('final_email_content', '')
        final_email_subject = session.get('final_email_subject', f"Procurement Inquiry - {task.task_name}")
        
        success_count = 0
        for supplier, assigned_items in selected_suppliers:
            assigned_item_ids = None
            quote_form_link = get_or_create_quote_form_link(task_id, supplier.id)

            if assigned_items:
                try:
                    assigned_item_ids = json.loads(assigned_items)
                except:
                    assigned_item_ids = None
            
            # Filter items for this supplier
            supplier_items = [item for item in pr_items if not assigned_item_ids or item.id in assigned_item_ids]
            
            sent_ok = send_procurement_email(
                supplier.email,
                supplier.name,
                supplier_items,
                task.task_name,
                assigned_item_ids,
                final_email_content,
                final_email_subject,
                supplier.contact_name,
                quote_form_link=quote_form_link
            )
            
            if sent_ok:
                success_count += 1
                # Update TaskSupplier
                task_supplier = db.session.query(TaskSupplier).filter_by(
                    task_id=task_id,
                    supplier_id=supplier.id
                ).first()
                
                if task_supplier and not task_supplier.initial_sent_at:
                    task_supplier.initial_sent_at = datetime.now()
                
                # Log email
                email_log = EmailLog(
                    task_id=task_id,
                    supplier_id=supplier.id,
                    email_type='initial',
                    subject=final_email_subject,
                    body=final_email_content,
                    status='sent'
                )
                db.session.add(email_log)
        
        # Update task status
        task.status = 'confirm_email'
        db.session.commit()
        
        # Clean up session data
        session.pop('email_content', None)
        session.pop('final_email_content', None)
        session.pop('final_email_subject', None)
        
        flash(f'Emails sent successfully! {success_count}/{len(selected_suppliers)} emails delivered.', 'success')
        return redirect(url_for('task_list'))
    
    return render_template('email_confirmation.html',
                         task=task,
                         suppliers=[s[0] for s in selected_suppliers],
                         pr_items=pr_items)

@app.route('/task-list')
def task_list():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    all_tasks = db.session.query(
        Task,
        User.username.label('created_by'),
        func.count(PRItem.id).label('item_count')  # Use func.count
    ).join(
        User, Task.user_id == User.id
    ).outerjoin(
        PRItem, Task.id == PRItem.task_id
    ).group_by(
        Task.id, User.username
    ).order_by(
        Task.created_at.desc()
    ).all()

    tasks_list = []
    for task, created_by, item_count in all_tasks:
        tasks_list.append({
            'id': task.id,
            'task_name': task.task_name,
            'user_id': task.user_id,
            'status': task.status,
            'created_at': task.created_at,
            'created_by': created_by,
            'item_count': item_count
        })

    return render_template('task_list.html', all_tasks=tasks_list)

@app.route('/task/<int:task_id>/follow-up', methods=['GET', 'POST'])
def follow_up(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))

    # Get suppliers with task supplier info
    suppliers_data = db.session.query(
        Supplier,
        TaskSupplier.assigned_items,
        TaskSupplier.initial_sent_at,
        TaskSupplier.followup_sent_at,
        TaskSupplier.replied_at
    ).join(
        TaskSupplier, Supplier.id == TaskSupplier.supplier_id
    ).filter(
        TaskSupplier.task_id == task_id,
        TaskSupplier.is_selected == True
    ).all()

    pr_items = db.session.query(PRItem).filter_by(task_id=task_id).all()

    # Suppliers eligible for follow-up: initial sent, not replied
    pending_suppliers = []
    for supplier, assigned_items, initial_sent_at, followup_sent_at, replied_at in suppliers_data:
        if initial_sent_at and not replied_at:
            pending_suppliers.append({
                'id': supplier.id,
                'name': supplier.name,
                'email': supplier.email,
                'contact_name': supplier.contact_name,
                'assigned_items': assigned_items,
                'initial_sent_at': initial_sent_at
            })

    default_body = session.get('followup_email_content') or """
    <p>Dear {supplier_name},</p>
    <p>This is a friendly follow-up regarding our procurement inquiry.</p>
    <p>Please share your quotation, lead time, and warranty terms at your earliest convenience.</p>
    <p>Thank you.</p>
    """
    default_subject = session.get('followup_email_subject') or f"Follow-up: Procurement Inquiry - {task.task_name}"

    if request.method == 'POST':
        body = request.form.get('email_content', default_body)
        subject = request.form.get('email_subject', default_subject)

        sent = 0
        for supplier_info in pending_suppliers:
            assigned_item_ids = None
            quote_form_link = get_or_create_quote_form_link(task_id, supplier_info['id'])

            if supplier_info['assigned_items']:
                try:
                    assigned_item_ids = json.loads(supplier_info['assigned_items'])
                except:
                    assigned_item_ids = None

            # Filter items
            supplier_items = [item for item in pr_items if not assigned_item_ids or item.id in assigned_item_ids]
            
            sent_ok = send_procurement_email(
                supplier_info['email'],
                supplier_info['name'],
                supplier_items,
                task.task_name,
                assigned_item_ids,
                body,
                subject,
                supplier_info.get('contact_name'),
                quote_form_link=quote_form_link
            )
            
            if sent_ok:
                sent += 1
                # Update followup sent time
                task_supplier = db.session.query(TaskSupplier).filter_by(
                    task_id=task_id,
                    supplier_id=supplier_info['id']
                ).first()
                if task_supplier:
                    task_supplier.followup_sent_at = datetime.now()
                
                # Log email
                email_log = EmailLog(
                    task_id=task_id,
                    supplier_id=supplier_info['id'],
                    email_type='followup',
                    subject=subject,
                    body=body,
                    status='sent'
                )
                db.session.add(email_log)
            else:
                # Log failed email
                email_log = EmailLog(
                    task_id=task_id,
                    supplier_id=supplier_info['id'],
                    email_type='followup',
                    subject=subject,
                    body=body,
                    status='failed',
                    error='send_failed'
                )
                db.session.add(email_log)
        
        db.session.commit()
        flash(f'Follow-up emails sent: {sent}/{len(pending_suppliers)}', 'success')
        return redirect(url_for('task_responses', task_id=task_id))

    return render_template('follow_up.html',
                       task=task,
                       pending_suppliers=pending_suppliers,
                       email_subject=default_subject,
                       email_content=default_body)

@app.route('/task/<int:task_id>/responses', methods=['GET', 'POST'])
def task_responses(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))

    if request.method == 'POST':
        action = request.form.get('action')
        supplier_id = request.form.get('supplier_id')
        if action == 'mark_replied' and supplier_id:
            task_supplier = db.session.query(TaskSupplier).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).first()
            if task_supplier:
                task_supplier.replied_at = datetime.now()
        elif action == 'mark_pending' and supplier_id:
            task_supplier = db.session.query(TaskSupplier).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).first()
            if task_supplier:
                task_supplier.replied_at = None
        db.session.commit()

    # Get suppliers with task info
    suppliers_data = db.session.query(
        Supplier,
        TaskSupplier.assigned_items,
        TaskSupplier.initial_sent_at,
        TaskSupplier.followup_sent_at,
        TaskSupplier.replied_at,
        TaskSupplier.quote_form_token
    ).join(
        TaskSupplier, Supplier.id == TaskSupplier.supplier_id
    ).filter(
        TaskSupplier.task_id == task_id,
        TaskSupplier.is_selected == True
    ).all()

    suppliers_list = []
    form_links = {}
    for supplier, assigned_items, initial_sent_at, followup_sent_at, replied_at, quote_form_token in suppliers_data:
        token = quote_form_token
        if not token:
            token = get_quote_serializer().dumps({'task_id': task_id, 'supplier_id': supplier.id})
            task_supplier = db.session.query(TaskSupplier).filter_by(
                task_id=task_id,
                supplier_id=supplier.id
            ).first()
            if task_supplier:
                task_supplier.quote_form_token = token
            db.session.commit()

        form_links[supplier.id] = public_url_for('supplier_quote_form', token=token)
        
        suppliers_list.append({
            'id': supplier.id,
            'name': supplier.name,
            'email': supplier.email,
            'contact_name': supplier.contact_name,
            'contact_number': supplier.contact_number,
            'assigned_items': assigned_items,
            'initial_sent_at': initial_sent_at,
            'followup_sent_at': followup_sent_at,
            'replied_at': replied_at,
            'quote_form_token': token
        })

    return render_template('responses.html', task=task, suppliers=suppliers_list, form_links=form_links)

def get_or_create_quote_form_link(task_id, supplier_id):
    """Get or create quote form token for a task-supplier pair."""
    
    task_supplier = db.session.query(TaskSupplier).filter_by(
        task_id=task_id,
        supplier_id=supplier_id
    ).first()

    if task_supplier and task_supplier.quote_form_token:
        token = task_supplier.quote_form_token
    else:
        token = get_quote_serializer().dumps({'task_id': task_id, 'supplier_id': supplier_id})
        
        if not task_supplier:
            # Create TaskSupplier record if it doesn't exist
            task_supplier = TaskSupplier(
                task_id=task_id,
                supplier_id=supplier_id,
                is_selected=True,
                quote_form_token=token
            )
            db.session.add(task_supplier)
        else:
            task_supplier.quote_form_token = token
        
        db.session.flush()

    return public_url_for("supplier_quote_form", token=token)

def get_optional_cert_file_id(request, task_id, supplier_id, pr_item_id, uid, old_cert_files):
    """
    COA is OPTIONAL:
    - If new file uploaded: save and return new id
    - Else: return old id (preserve previously uploaded file)
    - Else: None
    """
    file_key = f'cert_{uid}'
    cert_file_id = None

    cert_file = request.files.get(file_key)
    if cert_file and cert_file.filename:
        cert_file_id = save_uploaded_file(cert_file, task_id, supplier_id, pr_item_id)

    if not cert_file_id and pr_item_id in old_cert_files:
        cert_file_id = old_cert_files[pr_item_id]

    return cert_file_id


@app.route('/task/<int:task_id>/quotes/<int:supplier_id>', methods=['GET', 'POST'])
def capture_quotes(task_id, supplier_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    def form_last_nonempty(key: str):
        vals = request.form.getlist(key)
        for v in reversed(vals):
            if v is None:
                continue
            s = str(v).strip()
            if s != "":
                return s
        return None

    task = db.session.get(Task, task_id)
    supplier = db.session.get(Supplier, supplier_id)

    if not task or not supplier or (session['role'] != 'admin' and task.user_id != session['user_id']):
        flash('Task or supplier not found, or access denied.', 'error')
        return redirect(url_for('task_list'))

    pr_items = db.session.query(PRItem).filter_by(task_id=task_id).all()

    def load_existing_quotes():
        existing = db.session.query(SupplierQuote).filter_by(
            task_id=task_id,
            supplier_id=supplier_id
        ).all()
        
        quotes_map = {q.pr_item_id: q for q in existing}
        payment_terms_default = existing[0].payment_terms if existing else ''
        return quotes_map, payment_terms_default

    quotes_map, payment_terms_default = load_existing_quotes()

    # Get existing master quotation file ID
    master_quotation_file = None
    task_supplier = db.session.query(TaskSupplier).filter_by(
        task_id=task_id,
        supplier_id=supplier_id
    ).first()
    
    if task_supplier and task_supplier.quotation_file_id:
        master_quotation_file = task_supplier.quotation_file_id

    if request.method == 'POST':
        app.logger.info('capture_quotes POST received for task %s supplier %s', task_id, supplier_id)

        try:
            # 1) Preserve old certificate file IDs before deleting
            old_quotes = db.session.query(SupplierQuote).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).all()
            
            old_cert_files = {q.pr_item_id: q.cert_file_id for q in old_quotes}
            
            # Clear old quotes for this supplier+task
            db.session.query(SupplierQuote).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).delete()

            payment_terms_global = request.form.get('payment_terms') or None

            # Handle Master Quotation File
            quotation_file = request.files.get('quotation_file')
            if quotation_file and quotation_file.filename:
                # Save file (pr_item_id is None for general task files)
                q_file_id = save_uploaded_file(quotation_file, task_id, supplier_id, None)
                if q_file_id:
                    if not task_supplier:
                        task_supplier = TaskSupplier(
                            task_id=task_id,
                            supplier_id=supplier_id,
                            is_selected=True,
                            quotation_file_id=q_file_id
                        )
                        db.session.add(task_supplier)
                    else:
                        task_supplier.quotation_file_id = q_file_id

            # 2) insert new
            for item in pr_items:
                uid = str(item.id)
                app.logger.debug("Processing pr_item_id=%s uid=%s", item.id, uid)

                unit_price_val = parse_decimal_field(request.form.get(f'unit_price_{uid}') or None, f"Unit Price for Item {item.id}")
                unit_price = float(unit_price_val) if unit_price_val else None
                
                stock_availability = request.form.get(f'stock_availability_{uid}') or None
                lead_time = parse_optional_int(request.form.get(f'lead_time_{uid}') or None)
                warranty = request.form.get(f'warranty_{uid}') or None
                notes = request.form.get(f'notes_{uid}') or None
                ono = True if request.form.get(f"ono_{uid}") else False

                ono_width = parse_optional_int(request.form.get(f'ono_width_{uid}') or None)
                ono_length = parse_optional_int(request.form.get(f'ono_length_{uid}') or None)
                ono_thickness = parse_optional_int(request.form.get(f'ono_thickness_{uid}') or None)
                ono_dim_a = parse_optional_int(request.form.get(f'ono_dim_a_{uid}') or None)
                ono_dim_b = parse_optional_int(request.form.get(f'ono_dim_b_{uid}') or None)
                ono_diameter = parse_optional_int(request.form.get(f'ono_diameter_{uid}') or None)
                ono_uom = request.form.get(f'ono_uom_{uid}') or None
                ono_uom_qty = request.form.get(f'ono_uom_qty_{uid}') or None
                ono_brand = request.form.get(f'ono_brand_{uid}') or None

                cert_file_id = get_optional_cert_file_id(
                    request=request,
                    task_id=task_id,
                    supplier_id=supplier_id,
                    pr_item_id=item.id,
                    uid=uid,
                    old_cert_files=old_cert_files
                )

                has_any_input = any([
                    unit_price is not None, 
                    stock_availability, 
                    lead_time is not None, 
                    warranty, 
                    notes,
                    cert_file_id,
                    ono_width is not None, 
                    ono_length is not None, 
                    ono_thickness is not None,
                    ono_dim_a is not None, 
                    ono_dim_b is not None,
                    ono_diameter is not None,
                    ono_uom, 
                    ono_uom_qty, 
                    ono_brand
                ]) or ono

                if has_any_input:
                    quote = SupplierQuote(
                        task_id=task_id,
                        supplier_id=supplier_id,
                        pr_item_id=item.id,
                        unit_price=unit_price,
                        stock_availability=stock_availability,
                        cert_file_id=cert_file_id,
                        lead_time=lead_time,
                        warranty=warranty,
                        payment_terms=payment_terms_global,
                        ono=ono,
                        ono_width=ono_width,
                        ono_length=ono_length,
                        ono_thickness=ono_thickness,
                        ono_dim_a=ono_dim_a,
                        ono_dim_b=ono_dim_b,
                        ono_diameter=ono_diameter,
                        ono_uom=ono_uom,
                        ono_uom_qty=ono_uom_qty,
                        ono_brand=ono_brand,
                        notes=notes
                    )
                    db.session.add(quote)

            # 3) mark replied
            if not task_supplier:
                task_supplier = TaskSupplier(
                    task_id=task_id,
                    supplier_id=supplier_id,
                    is_selected=True,
                    replied_at=datetime.now()
                )
                db.session.add(task_supplier)
            else:
                task_supplier.replied_at = datetime.now()

            db.session.commit()
            flash('Quotes saved.', 'success')
            return redirect(url_for('task_responses', task_id=task_id))

        except Exception as e:
            db.session.rollback()
            app.logger.exception("Error while saving quotes (task_id=%s supplier_id=%s): %s", task_id, supplier_id, str(e))
            flash(f"We couldn't save the quotes. Error: {str(e)}", "error")
            
            # If we reach here: POST failed, re-load what's currently in DB
            quotes_map, payment_terms_default = load_existing_quotes()

    # GET or failed POST -> show form
    # Convert quotes_map to dict format for template
    quotes_map_dict = {}
    for pr_item_id, quote in quotes_map.items():
        quotes_map_dict[pr_item_id] = {
            'id': quote.id,
            'task_id': quote.task_id,
            'supplier_id': quote.supplier_id,
            'pr_item_id': quote.pr_item_id,
            'unit_price': quote.unit_price,
            'stock_availability': quote.stock_availability,
            'lead_time': quote.lead_time,
            'warranty': quote.warranty,
            'payment_terms': quote.payment_terms,
            'notes': quote.notes,
            'ono': quote.ono,
            'ono_width': quote.ono_width,
            'ono_length': quote.ono_length,
            'ono_thickness': quote.ono_thickness,
            'ono_dim_a': quote.ono_dim_a,
            'ono_dim_b': quote.ono_dim_b,
            'ono_diameter': quote.ono_diameter,
            'ono_uom': quote.ono_uom,
            'ono_uom_qty': quote.ono_uom_qty,
            'ono_brand': quote.ono_brand,
            'cert_file_id': quote.cert_file_id
        }

    return render_template(
        'quotes_form.html',
        task=task,
        supplier=supplier,
        pr_items=pr_items,
        quotes_map=quotes_map_dict,
        payment_terms_default=payment_terms_default,
        master_quotation_file=master_quotation_file
    )

@app.route('/supplier/quote-form/<token>', methods=['GET', 'POST'])
def supplier_quote_form(token):
    try:
        data = get_quote_serializer().loads(token)
        task_id = data.get('task_id')
        supplier_id = data.get('supplier_id')
    except BadSignature:
        return "Invalid or expired link", 400

    # Use db.session instead of get_db_connection()
    task = db.session.get(Task, task_id)
    supplier = db.session.get(Supplier, supplier_id)
    
    if not task or not supplier:
        return "Task or supplier not found", 404

    def form_last_nonempty(key: str):
        vals = request.form.getlist(key)
        for v in reversed(vals):
            if v is None:
                continue
            s = str(v).strip()
            if s != "":
                return s
        return None

    # Get TaskSupplier using SQLAlchemy
    task_supplier = db.session.query(TaskSupplier).filter_by(
        task_id=task_id,
        supplier_id=supplier_id
    ).first()

    assigned_ids = None
    if task_supplier and task_supplier.assigned_items:
        try:
            assigned_ids = [int(x) for x in json.loads(task_supplier.assigned_items)]
        except Exception:
            assigned_ids = None

    # Get PR items using SQLAlchemy
    pr_items_query = db.session.query(PRItem).filter_by(task_id=task_id)
    
    if assigned_ids:
        pr_items = pr_items_query.filter(PRItem.id.in_(assigned_ids)).all()
    else:
        pr_items = pr_items_query.all()

    if request.method == 'POST':
        try:
            # 1) Preserve old certificate file IDs before deleting
            old_quotes = db.session.query(SupplierQuote).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).all()
            old_cert_files = {q.pr_item_id: q.cert_file_id for q in old_quotes}
            
            # Clear old quotes for this supplier+task
            db.session.query(SupplierQuote).filter_by(
                task_id=task_id,
                supplier_id=supplier_id
            ).delete()

            payment_terms_global = request.form.get('payment_terms') or None

            # Handle Master Quotation File
            quotation_file = request.files.get('quotation_file')
            if quotation_file and quotation_file.filename:
                # Save file (pr_item_id is None for general task files)
                q_file_id = save_uploaded_file(quotation_file, task_id, supplier_id, None)
                if q_file_id:
                    if not task_supplier:
                        task_supplier = TaskSupplier(
                            task_id=task_id,
                            supplier_id=supplier_id,
                            is_selected=True,
                            quotation_file_id=q_file_id
                        )
                        db.session.add(task_supplier)
                    else:
                        task_supplier.quotation_file_id = q_file_id

            # 2) Insert new quotes (similar to your capture_quotes function)
            for item in pr_items:
                uid = str(item.id)

                unit_price_val = parse_decimal_field(request.form.get(f'unit_price_{uid}') or None, f"Unit Price for Item {item.id}")
                unit_price = float(unit_price_val) if unit_price_val else None
                
                stock_availability = request.form.get(f'stock_availability_{uid}') or None
                lead_time = parse_optional_int(request.form.get(f'lead_time_{uid}') or None)
                warranty = request.form.get(f'warranty_{uid}') or None
                notes = request.form.get(f'notes_{uid}') or None
                ono = 1 if request.form.get(f"ono_{uid}") else 0

                # O.N.O. alternate dimensions
                ono_width = parse_optional_int(request.form.get(f'ono_width_{uid}') or None)
                ono_length = parse_optional_int(request.form.get(f'ono_length_{uid}') or None)
                ono_thickness = parse_optional_int(request.form.get(f'ono_thickness_{uid}') or None)
                ono_dim_a = parse_optional_int(request.form.get(f'ono_dim_a_{uid}') or None)
                ono_dim_b = parse_optional_int(request.form.get(f'ono_dim_b_{uid}') or None)
                ono_diameter = parse_optional_int(request.form.get(f'ono_diameter_{uid}') or None)
                ono_uom = request.form.get(f'ono_uom_{uid}') or None
                ono_uom_qty = request.form.get(f'ono_uom_qty_{uid}') or None
                ono_brand = request.form.get(f'ono_brand_{uid}') or None

                cert_file_id = get_optional_cert_file_id(
                    request=request,
                    task_id=task_id,
                    supplier_id=supplier_id,
                    pr_item_id=item.id,
                    uid=uid,
                    old_cert_files=old_cert_files
                )

                # Save row if ANY meaningful input exists, OR ONO checked
                has_any_input = any([
                    unit_price, stock_availability, lead_time, warranty, notes,
                    cert_file_id,
                    ono_width, ono_length, ono_thickness,
                    ono_dim_a, ono_dim_b, ono_diameter,
                    ono_uom, ono_uom_qty, ono_brand
                ]) or (ono == 1)

                if has_any_input:
                    quote = SupplierQuote(
                        task_id=task_id,
                        supplier_id=supplier_id,
                        pr_item_id=item.id,
                        unit_price=unit_price,
                        stock_availability=stock_availability,
                        cert_file_id=cert_file_id,
                        lead_time=lead_time,
                        warranty=warranty,
                        payment_terms=payment_terms_global,
                        ono=bool(ono),
                        ono_width=ono_width,
                        ono_length=ono_length,
                        ono_thickness=ono_thickness,
                        ono_dim_a=ono_dim_a,
                        ono_dim_b=ono_dim_b,
                        ono_diameter=ono_diameter,
                        ono_uom=ono_uom,
                        ono_uom_qty=ono_uom_qty,
                        ono_brand=ono_brand,
                        notes=notes
                    )
                    db.session.add(quote)

            # 3) Mark replied + log
            if not task_supplier:
                task_supplier = TaskSupplier(
                    task_id=task_id,
                    supplier_id=supplier_id,
                    is_selected=True
                )
                db.session.add(task_supplier)
            
            task_supplier.replied_at = datetime.now()

            # Log email
            email_log = EmailLog(
                task_id=task_id,
                supplier_id=supplier_id,
                email_type='supplier_form',
                subject=f'Quote submitted by {supplier.name}',
                body=None,
                status='received'
            )
            db.session.add(email_log)

            db.session.commit()
            return render_template('supplier_form_success.html', supplier=supplier, task=task)

        except Exception as e:
            db.session.rollback()
            app.logger.exception("Error while saving supplier form (task_id=%s supplier_id=%s): %s", task_id, supplier_id, str(e))
            flash("We couldn't save your quotation due to a system issue. Please try again.", "error")
            # Fall through to re-render the form

    # GET or failed POST -> show form again
    return render_template(
        'supplier_public_quote.html',
        task=task,
        supplier=supplier,
        pr_items=pr_items
    )

@app.route('/debug/quotes/<int:task_id>/<int:supplier_id>', methods=['GET'])
def debug_quotes(task_id, supplier_id):
    """Return JSON dump of supplier_quotes for debugging (task+supplier)."""
    if 'user_id' not in session:
        return jsonify({'error': 'login required'}), 403

    quotes = db.session.query(SupplierQuote).filter_by(
        task_id=task_id,
        supplier_id=supplier_id
    ).all()
    
    quotes_list = []
    for quote in quotes:
        quotes_list.append({
            'id': quote.id,
            'task_id': quote.task_id,
            'supplier_id': quote.supplier_id,
            'pr_item_id': quote.pr_item_id,
            'unit_price': float(quote.unit_price) if quote.unit_price else None,
            'stock_availability': quote.stock_availability,
            'lead_time': quote.lead_time,
            'warranty': quote.warranty,
            'payment_terms': quote.payment_terms,
            'notes': quote.notes,
            'ono': quote.ono,
            'ono_width': quote.ono_width,
            'ono_length': quote.ono_length,
            'ono_thickness': quote.ono_thickness,
            'ono_dim_a': quote.ono_dim_a,
            'ono_dim_b': quote.ono_dim_b,
            'ono_diameter': quote.ono_diameter,
            'ono_uom': quote.ono_uom,
            'ono_uom_qty': quote.ono_uom_qty,
            'ono_brand': quote.ono_brand,
            'cert_file_id': quote.cert_file_id
        })
    
    return jsonify(quotes_list)

@app.route('/task/<int:task_id>/export-comparison')
def export_comparison(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))

    # Get PR items
    pr_items_result = db.session.query(PRItem).filter_by(task_id=task_id).all()
    pr_items = []
    for item in pr_items_result:
        pr_items.append({
            'id': item.id,
            'task_id': item.task_id,
            'item_category': item.item_category,
            'item_name': item.item_name,
            'brand': item.brand,
            'quantity': item.quantity,
            'payment_terms': item.payment_terms,
            'width': item.width,
            'length': item.length,
            'thickness': item.thickness,
            'dim_a': item.dim_a,
            'dim_b': item.dim_b,
            'diameter': item.diameter,
            'uom_qty': item.uom_qty,
            'uom': item.uom,
            'our_remarks': item.our_remarks
        })
    
    # Get quotes with supplier info
    quotes_result = db.session.query(
        SupplierQuote,
        Supplier.name.label('supplier_name'),
        TaskSupplier.replied_at,
        TaskSupplier.quotation_file_id
    ).join(
        Supplier, SupplierQuote.supplier_id == Supplier.id
    ).outerjoin(
        TaskSupplier, 
        db.and_(
            TaskSupplier.supplier_id == SupplierQuote.supplier_id,
            TaskSupplier.task_id == SupplierQuote.task_id
        )
    ).filter(
        SupplierQuote.task_id == task_id
    ).all()
    
    quotes = []
    for quote, supplier_name, replied_at, quotation_file_id in quotes_result:
        quotes.append({
            'id': quote.id,
            'task_id': quote.task_id,
            'supplier_id': quote.supplier_id,
            'pr_item_id': quote.pr_item_id,
            'unit_price': float(quote.unit_price) if quote.unit_price else None,
            'stock_availability': quote.stock_availability,
            'lead_time': quote.lead_time,
            'warranty': quote.warranty,
            'payment_terms': quote.payment_terms,
            'notes': quote.notes,
            'ono': quote.ono,
            'ono_width': quote.ono_width,
            'ono_length': quote.ono_length,
            'ono_thickness': quote.ono_thickness,
            'ono_dim_a': quote.ono_dim_a,
            'ono_dim_b': quote.ono_dim_b,
            'ono_diameter': quote.ono_diameter,
            'ono_uom': quote.ono_uom,
            'ono_uom_qty': quote.ono_uom_qty,
            'ono_brand': quote.ono_brand,
            'cert_file_id': quote.cert_file_id,
            'supplier_name': supplier_name,
            'replied_at': replied_at,
            'quotation_file_id': quotation_file_id
        })
    
    # Define dimension spec function early
    def get_dim_spec(cat: str):
        cat = (cat or "").strip()
        if cat in ["Steel Plates", "Stainless Steel"]:
            return ["W (mm)", "L (mm)", "Thk (mm)"], 3
        if cat == "Angle Bar":
            return ["A (mm)", "B (mm)", "L (mm)", "Thk (mm)"], 4
        if cat in ["Rebar", "Bolts, Fasteners"]:
            return ["D (mm)", "L (mm)"], 2
        # Other categories - single "Packing" column (no sub-header needed)
        return [""], 1

    cat = (pr_items[0].get("item_category") or "").strip() if pr_items else ""

    WEIGHT_CATS = {"Steel Plates", "Stainless Steel", "Angle Bar", "Rebar", "Bolts, Fasteners"}
    has_weight = cat in WEIGHT_CATS

    dim_group_label = "Dimensions" if has_weight else "Packing"
    dim_headers, DIM_COUNT = get_dim_spec(cat)

    # BASE columns:
    # 3 fixed + DIM_COUNT dims + Qty + (Weight only if has_weight)
    BASE_COLS = 3 + DIM_COUNT + 1 + (1 if has_weight else 0)
    SUPPLIER_START_COL = BASE_COLS + 1

    def to_float(x):
        try:
            if x is None or x == "":
                return None
            return float(x)
        except (ValueError, TypeError):
            return None

    def lead_time_to_days(v):
        """Convert common lead time text into comparable 'days' number."""
        if v is None:
            return None
        s = str(v).strip().lower()
        if not s:
            return None

        # treat ready stock / immediate as 0 days (best)
        if "ready" in s or "immediate" in s or "instock" in s or "in stock" in s:
            return 0.0

        m = re.search(r"(\d+(\.\d+)?)", s)
        if not m:
            return None
        num = float(m.group(1))

        if "week" in s:
            return num * 7.0
        if "month" in s:
            return num * 30.0
        if "day" in s:
            return num

        # If user typed plain "7" assume days
        return num

    def _norm(x):
        """Normalize dimension cell value to a clean string for Excel + grouping."""
        if x is None:
            return ""
        s = str(x).strip()
        return "" if s.lower() in ("none", "null") else s

    def _num(x):
        """Convert dimension to float if possible."""
        try:
            s = _norm(x)
            return None if s == "" else float(s)
        except Exception:
            return None

    def _is_filled(v) -> bool:
        if v is None:
            return False
        s = str(v).strip()
        return s != "" and s.lower() not in ("none", "null")

    def has_any_ono_value(category: str, q: dict) -> bool:
        category = (category or "").strip()

        if category in ["Steel Plates", "Stainless Steel"]:
            keys = ["ono_width", "ono_length", "ono_thickness", "ono_brand"]
        elif category == "Angle Bar":
            keys = ["ono_dim_a", "ono_dim_b", "ono_length", "ono_thickness", "ono_brand"]
        elif category in ["Rebar", "Bolts, Fasteners"]:
            keys = ["ono_diameter", "ono_length", "ono_brand"]
        else:
            keys = ["ono_uom_qty", "ono_uom", "ono_brand"]

        return any(_is_filled(q.get(k)) for k in keys)

    def dims_for_row(category: str, item: dict, q: dict = None, use_ono: bool = False):
        """
        Returns dims (length depends on category).
        If use_ono is True, it uses ONO fields where filled, otherwise falls back to item dims.
        """
        category = (category or "").strip()

        def pick(ono_key, item_key):
            # Prefer ONO only when enabled AND field is filled
            if use_ono and q and _is_filled(q.get(ono_key)):
                return str(q.get(ono_key)).strip()

            v = item.get(item_key)
            return "" if v is None else str(v).strip()

        if category in ["Steel Plates", "Stainless Steel"]:
            return [
                pick("ono_width", "width"),
                pick("ono_length", "length"),
                pick("ono_thickness", "thickness"),
            ]

        if category == "Angle Bar":
            return [
                pick("ono_dim_a", "dim_a"),
                pick("ono_dim_b", "dim_b"),
                pick("ono_length", "length"),
                pick("ono_thickness", "thickness"),
            ]

        if category in ["Rebar", "Bolts, Fasteners"]:
            return [
                pick("ono_diameter", "diameter"),
                pick("ono_length", "length"),
            ]

        return [
            pick("ono_uom_qty", "uom_qty"),
        ]

    def weight_for_dims(category: str, dims: list, base_weight):
        """
        Compute weight for the CURRENT row dims (incl ONO dims).
        If cannot compute, return base_weight.
        """
        category = (category or "").strip()

        # Helpers
        def _num(x):
            try:
                if x is None:
                    return None
                s = str(x).strip()
                if s == "" or s.lower() in ("none", "null"):
                    return None
                return float(s)
            except Exception:
                return None

        # Steel Plates / Stainless Steel: (W * L * Thk * 7.85) / 1,000,000
        if category in ["Steel Plates", "Stainless Steel"]:
            # dims = [W, L, Thk]
            w = _num(dims[0])
            L = _num(dims[1])
            thk = _num(dims[2])
            if w is not None and L is not None and thk is not None:
                return round((w * L * thk * 7.85) / 1_000_000, 2)
            return base_weight

        # Rebar / Bolts: (π/4) * D² * L * 7.85 * 10⁻⁶
        if category in ["Rebar", "Bolts, Fasteners"]:
            # dims = [D, L]
            d = _num(dims[0])
            L = _num(dims[1])
            if d is not None and L is not None:
                return round((math.pi / 4) * (d ** 2) * L * 7.85 * 0.000001, 2)
            return base_weight

        # Angle Bar: (L * Thk * (A + B - Thk) * 7.85) / 1,000,000
        if category == "Angle Bar":
            # dims = [A, B, L, Thk]
            a = _num(dims[0])
            b = _num(dims[1])
            L = _num(dims[2])
            thk = _num(dims[3])
            if a is not None and b is not None and L is not None and thk is not None:
                return round((L * thk * (a + b - thk) * 7.85) / 1_000_000, 2)
            return base_weight

        # Other categories
        return base_weight

    SUPPLIER_HEADERS = [
        "Quoted Price (RM)",
        "Total Amount Quoted (RM)",
        "Delivery Lead Time (Days)",
        "Stock Availability",
        "COA",
        "Remarks",
    ]

    # Only add Rate if we have weight
    if has_weight:
        SUPPLIER_HEADERS = ["Rate (RM/Kg)"] + SUPPLIER_HEADERS

    SUPPLIER_BLOCK_COLS = len(SUPPLIER_HEADERS)

    # Offsets must be computed AFTER SUPPLIER_HEADERS is finalized
    def idx(name): 
        return SUPPLIER_HEADERS.index(name)

    OFF_RATE   = idx("Rate (RM/Kg)") if has_weight else None
    OFF_PRICE  = idx("Quoted Price (RM)")
    OFF_TOTAL  = idx("Total Amount Quoted (RM)")
    OFF_LEAD   = idx("Delivery Lead Time (Days)")
    OFF_STOCK  = idx("Stock Availability")
    OFF_COA    = idx("COA")
    OFF_REMARKS = idx("Remarks")

    suppliers = {}  # supplier_id -> supplier_name
    supplier_terms = {}  # supplier_id -> set(payment_terms)
    supplier_quotation_files = {}  # supplier_id -> quotation_file_id

    for q in quotes:
        sid = q['supplier_id']
        suppliers.setdefault(sid, q['supplier_name'])

        pt = (q['payment_terms'] or "").strip()
        if pt:
            supplier_terms.setdefault(sid, set()).add(pt)
        
        # Store quotation file ID if available
        if q['quotation_file_id']:
            supplier_quotation_files[sid] = q['quotation_file_id']

    suppliers_list = sorted(suppliers.items(), key=lambda x: x[1])  # (sid, name)

    def supplier_header_label(supplier_id, supplier_name):
        terms = supplier_terms.get(supplier_id, set())
        terms_str = ""
        if terms:
            if len(terms) == 1:
                terms_str = f" ({next(iter(terms))})"
            else:
                terms_str = " (Multiple)"
        
        label = supplier_name + terms_str
        return label

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    row1 = [
        "Item Name (O.N.O.)",
        "Brand/Specification",
        "Category",
    ] + ([dim_group_label] + [""] * (DIM_COUNT - 1)) + [
        "Qty",
    ]

    if has_weight:
        row1.append("Weight (Kg)")

    for supplier_id, supplier_name in suppliers_list:
        row1.extend([supplier_header_label(supplier_id, supplier_name)] + [""] * (SUPPLIER_BLOCK_COLS - 1))
    ws.append(row1)

    row2 = ["", "", ""] + dim_headers + [""]  # blank under Qty
    if has_weight:
        row2.append("")  # blank under Weight

    for supplier_id, supplier_name in suppliers_list:
        row2.extend(SUPPLIER_HEADERS)
    ws.append(row2)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # ---- Header styling (Row 1–2) ----
    header_fill = PatternFill(fill_type="solid", fgColor="FF2b86e1")   # ARGB - Hercules Blue
    header_font = Font(bold=True, color="FFFFFFFF")                   # WHITE, visible
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    best_fill  = PatternFill(fill_type="solid", fgColor="FFFFEB9C")   # ARGB
    total_fill = PatternFill(fill_type="solid", fgColor="FFD9D9D9")   # ARGB

    # Make sure we color all header cells across the full table width
    max_col = ws.max_column

    for r in (1, 2):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    # Merge and center Item Name, Brand, Category (cols 1-3)
    for col in range(1, 4):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center

    # Merge and center the dim group header
    dim_start = 4
    dim_end = 3 + DIM_COUNT

    if has_weight:
        # Normal case: Row 1 group header + Row 2 subheaders (W/L/Thk etc.)
        ws.merge_cells(start_row=1, start_column=dim_start, end_row=1, end_column=dim_end)
        cell = ws.cell(row=1, column=dim_start)
        cell.value = dim_group_label  # "Dimensions"
        cell.font = bold
        cell.alignment = center

        # Center dim subheaders in row2
        for col in range(dim_start, dim_end + 1):
            c = ws.cell(row=2, column=col)
            c.font = bold
            c.alignment = center
    else:
        # "Other" categories: Packing has NO second header
        # Merge Row 1–2 for Packing
        ws.merge_cells(start_row=1, start_column=dim_start, end_row=2, end_column=dim_end)
        cell = ws.cell(row=1, column=dim_start)
        cell.value = dim_group_label  # "Packing"
        cell.font = bold
        cell.alignment = center

    # Center dim headers in row2
    for col in range(dim_start, dim_end + 1):
        c = ws.cell(row=2, column=col)
        c.font = bold
        c.alignment = center

    # Merge Quantity and Weight columns (dynamic positions)
    qty_col = dim_end + 1
    wt_col  = dim_end + 2  # only valid if has_weight

    # Always merge Qty (row 1-2)
    ws.merge_cells(start_row=1, start_column=qty_col, end_row=2, end_column=qty_col)
    c = ws.cell(row=1, column=qty_col)
    c.font = bold
    c.alignment = center

    # Merge Weight only when applicable
    if has_weight:
        ws.merge_cells(start_row=1, start_column=wt_col, end_row=2, end_column=wt_col)
        c = ws.cell(row=1, column=wt_col)
        c.font = bold
        c.alignment = center

    # Merge and center each supplier header (8 columns each)
    col_idx = SUPPLIER_START_COL
    for supplier_id, supplier_name in suppliers_list:
        ws.merge_cells(
            start_row=1, start_column=col_idx,
            end_row=1, end_column=col_idx + (SUPPLIER_BLOCK_COLS - 1)
        )
        c = ws.cell(row=1, column=col_idx)
        c.alignment = center

        qfile_id = supplier_quotation_files.get(supplier_id)
        if qfile_id:
            mq_url = public_url_for("serve_file", file_id=qfile_id)
            c.hyperlink = mq_url
            # hyperlink look - use Hercules blue
            c.font = Font(bold=True, color="FF003366", underline="single")  # Hercules Blue
        else:
            # normal look
            c.font = Font(bold=True, color="FF000000")  # BLACK

        col_idx += SUPPLIER_BLOCK_COLS

    # Format row 2 as bold and centered
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FF000000")
        cell.alignment = center

    best_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # light gray

    quotes_by_item = {}
    for q in quotes:
        item_id = q['pr_item_id']
        quotes_by_item.setdefault(item_id, {})
        quotes_by_item[item_id][q['supplier_id']] = dict(q)

    cert_links = {}  # (row, col) -> url
    supplier_total_amounts = {sid: 0.0 for sid, _ in suppliers_list}

    current_row = 3
    for item in pr_items:
        category = (item['item_category'] or '').strip()
        w = item['width'] or None
        l = item['length'] or None
        thk = item['thickness'] or None
        dim_a = item['dim_a'] or None
        dim_b = item['dim_b'] or None
        diameter = item['diameter'] or None
        uom_qty = item['uom_qty'] or None
        uom = item['uom'] or None

        # Calculate weight based on category
        weight = ""
        weight_formula_applied = False
        
        # Steel Plates / Stainless Steel: (W * L * Thk * 7.85) / 1,000,000
        if category in ['Steel Plates', 'Stainless Steel']:
            if w and l and thk:
                try:
                    w_val = float(w)
                    l_val = float(l)
                    thk_val = float(thk)
                    weight = round((w_val * l_val * thk_val * 7.85) / 1000000, 2)
                    weight_formula_applied = True
                except (ValueError, TypeError):
                    weight = ""
        
        # Rebar / Bolts, Fasteners: (π/4) * D² * L * 7.85 * 10⁻⁶
        elif category in ['Rebar', 'Bolts, Fasteners']:
            if diameter and l:
                try:
                    d_val = float(diameter)
                    l_val = float(l)
                    weight = round((math.pi / 4) * (d_val ** 2) * l_val * 7.85 * 0.000001, 2)
                    weight_formula_applied = True
                except (ValueError, TypeError):
                    weight = ""
        
        # Angle Bar: (L * Thk * (A + B - Thk) * 7.85) / 1,000,000
        elif category == 'Angle Bar':
            if l and thk and dim_a and dim_b:
                try:
                    l_val = float(l)
                    thk_val = float(thk)
                    a_val = float(dim_a)
                    b_val = float(dim_b)
                    weight = round((l_val * thk_val * (a_val + b_val - thk_val) * 7.85) / 1000000, 2)
                    weight_formula_applied = True
                except (ValueError, TypeError):
                    weight = ""
        
        # For "Other" categories, no weight calculation
        else:
            weight = "N/A"

        # Build dimension display for this item
        # dim_display order = [Dim1, Dim2, Dim3, Dim4]
        # Dim1 = W/A/D/UOMQty, Dim2 = B, Dim3 = L, Dim4 = Thk
        base_dim_display = ["", "", "", ""]
        if category in ["Steel Plates", "Stainless Steel"]:
            w = item.get('width') or ''
            l = item.get('length') or ''
            thk = item.get('thickness') or ''
            base_dim_display = [w, "", l, thk]
        elif category == "Angle Bar":
            a = item.get('dim_a') or ''
            b = item.get('dim_b') or ''
            l = item.get('length') or ''
            thk = item.get('thickness') or ''
            base_dim_display = [a, b, l, thk]
        elif category in ["Rebar", "Bolts, Fasteners"]:
            d = item.get('diameter') or ''
            l = item.get('length') or ''
            base_dim_display = [d, "", l, ""]
        else:
            uom = item.get('uom') or ''
            uom_qty = item.get('uom_qty') or ''
            base_dim_display = [uom_qty, "", "", ""]

        item_quotes_map = quotes_by_item.get(item['id'], {})
        
        metric_candidates = {
            "rate": [],
            "price": [],
            "total": [],
            "lead": []
        }

        # ---------------------------------------------------------
        # Group quotes by "Dimension Key"
        # Key: (is_ono, (dim1, dim2, dim3, dim4))
        #   - Standard Quote: (False, base_dim_display)
        #   - O.N.O Quote:    (True, (d1, d2, d3... from ono cols))
        # ---------------------------------------------------------
        
        # Structure: key -> { supplier_id: quote }
        grouped_quotes = {}
        
        # Ensure the "Standard" group always exists
        standard_key = (False, tuple(base_dim_display))
        grouped_quotes[standard_key] = {} 
        
        # Distribute quotes into groups
        # Normalize standard dims so grouping is consistent
        base_dim_display = dims_for_row(category, item, q=None)  # correct length
        standard_dims = tuple(_norm(x) for x in base_dim_display)

        item_brand_norm = _norm(item.get("brand") or "")
        standard_key = (False, standard_dims, item_brand_norm)
        grouped_quotes = {standard_key: {}}

        # Distribute quotes into groups (STANDARD vs ONO with fallback + ONO brand)
        for supplier_id, supplier_name in suppliers_list:
            q = item_quotes_map.get(supplier_id)
            if not q:
                continue

            use_ono = has_any_ono_value(category, q)

            eff_dims = dims_for_row(category, item, q=q, use_ono=use_ono)
            eff_dims_norm = tuple(_norm(x) for x in eff_dims)

            # ✅ ONO brand override: if provided, we split into its own ONO row
            ono_brand_val = _norm(q.get("ono_brand")) if use_ono and _is_filled(q.get("ono_brand")) else ""

            if use_ono:
                # If supplier sets ono_brand, use it; otherwise fall back to the item brand
                eff_brand = ono_brand_val or item_brand_norm
                key = (True, eff_dims_norm, eff_brand)
                grouped_quotes.setdefault(key, {})[supplier_id] = q
            else:
                grouped_quotes[standard_key][supplier_id] = q


        # ---------------------------------------------------------
        # Render Rows
        # ---------------------------------------------------------
        def sort_keys(k):
            is_ono, dims, brand_val = k
            return (is_ono, dims, brand_val)

        sorted_keys = sorted(grouped_quotes.keys(), key=sort_keys)
        
        for key in sorted_keys:
            is_ono, dims, brand_val = key
            quotes_for_row_map = grouped_quotes[key]
            
            # Skip empty O.N.O rows
            if is_ono and not quotes_for_row_map:
                continue

            # Construct Label
            row_label = f"{item['item_name']} (O.N.O)" if is_ono else item['item_name']
            base_weight = item.get("weight")  # if you have it in DB, else None
            row_weight = weight_for_dims(category, list(dims), base_weight)

            # Determine Quantity display (append UOM for "Other" categories)
            qty_display = item["quantity"] or ""
            if not has_weight and qty_display:
                # For non-weight categories (Other), usually implies UOM-based
                # Get effective UOM for this row
                if not is_ono:
                    eff_uom = item.get("uom")
                else:
                    # Pick UOM from any quote in this group
                    # (Assuming grouped by Packing size implicitly groups by UOM or UOM is consistent)
                    any_q = next(iter(quotes_for_row_map.values()))
                    eff_uom = any_q.get("ono_uom")
                
                if eff_uom:
                    qty_display = f"{qty_display} {eff_uom}"

            row = [
                row_label,
                brand_val or (item["brand"] or ""),
                category,
                *dims,
                qty_display,
            ]

            row_weight = None
            if has_weight:
                base_weight = item.get("weight")
                row_weight = weight_for_dims(category, list(dims), base_weight)
                row.append(row_weight)
            
            # Fill supplier columns
            # Metric candidates for highlighting *within this row*
            row_metric_candidates = {
                "price": [],
                "total": [],
                "lead": []
            }
            if has_weight:
                row_metric_candidates["rate"] = []

            col_idx = SUPPLIER_START_COL
            for supplier_id, supplier_name in suppliers_list:
                q = quotes_for_row_map.get(supplier_id)
                
                # Logic: Quotes are already pre-filtered by the group key
                # So we just render 'q' if it exists.
                
                should_include = True # effectively always true if q exists in this map
                if q:
                    # q is guaranteed to match the row type thanks to grouping logic
                    pass
                
                if q and should_include:
                    cert_display = ""
                    cert_url = None
                    
                    # Handle File Asset ID (New) or Legacy Path
                    if q.get('cert_file_id'):
                        cert_display = "View File"
                        # Generate full URL pointing to /file/<id>
                        cert_url = public_url_for("serve_file", file_id=q["cert_file_id"])
                    elif q.get('cert'):
                        # Legacy fallback
                        val = str(q['cert'])
                        if val.isdigit():
                            cert_display = "View File"
                            cert_url = public_url_for("serve_file", file_id=int(val))
                        else:
                            # Old path style
                            cert_display = "View File"
                            # Note: Local paths won't open in Excel, but preserving legacy behavior
                            cert_url = val 

                    unit_price = q['unit_price'] if q['unit_price'] is not None else ""
                    unit_price_val = to_float(unit_price)

                    qty_val = to_float(item['quantity'])

                    rate_val = ""
                    if has_weight:
                        weight_val = to_float(row_weight)
                        if unit_price_val is not None and weight_val not in (None, 0):
                            rate_val = round(unit_price_val / weight_val, 4)

                    total_amount_val = ""
                    if unit_price_val is not None and qty_val is not None:
                        total_amount_val = round(unit_price_val * qty_val, 2)
                        supplier_total_amounts[supplier_id] += float(total_amount_val)

                    # Build supplier cells in the same order as SUPPLIER_HEADERS
                    supplier_cells = []

                    if has_weight:
                        supplier_cells.append(rate_val)

                    supplier_cells.extend([
                        unit_price,
                        total_amount_val,
                        q.get('lead_time') or "",
                        q.get('stock_availability') or "",
                        cert_display,
                        q.get('notes') or ""
                    ])

                    row.extend(supplier_cells)

                    if cert_url:
                        cert_links[(current_row, col_idx + OFF_COA)] = cert_url

                    # highlight candidates
                    if has_weight:
                        row_metric_candidates["rate"].append((col_idx + OFF_RATE, to_float(rate_val)))

                    row_metric_candidates["price"].append((col_idx + OFF_PRICE, unit_price_val))
                    row_metric_candidates["total"].append((col_idx + OFF_TOTAL, to_float(total_amount_val)))
                    row_metric_candidates["lead"].append((col_idx + OFF_LEAD, lead_time_to_days(q.get('lead_time'))))

                    col_idx += SUPPLIER_BLOCK_COLS
                else:
                    # No quote from this supplier OR quote belongs to the other row type
                    row.extend([""] * SUPPLIER_BLOCK_COLS)
                    col_idx += SUPPLIER_BLOCK_COLS

            ws.append(row)

            # Apply per-row best highlighting (lowest wins)
            for k in (["rate", "price", "total", "lead"] if has_weight else ["price", "total", "lead"]):
                vals = [(col, v) for (col, v) in row_metric_candidates[k] if v is not None]
                if not vals:
                    continue
                best_val = min(v for _, v in vals)
                for col, v in vals:
                    if v == best_val:
                        ws.cell(row=current_row, column=col).fill = best_fill

            current_row += 1 
    
    # After rows loop
    totals_row = current_row
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=BASE_COLS)
    label_cell = ws.cell(row=totals_row, column=1)
    label_cell.value = "TOTAL (RM)"
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.fill = total_fill

    total_cells = []  # (col, value)
    for i, (sid, _sname) in enumerate(suppliers_list):
        total_col = SUPPLIER_START_COL + i * SUPPLIER_BLOCK_COLS + OFF_TOTAL
        c = ws.cell(row=totals_row, column=total_col)
        val = supplier_total_amounts.get(sid, 0.0)

        if val and val != 0.0:
            c.value = round(val, 2)
            c.number_format = '#,##0.00'
            total_cells.append((total_col, float(c.value)))
        else:
            c.value = ""

        c.font = Font(bold=True)
        c.fill = total_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight best (lowest) total
    if total_cells:
        best_total = min(v for _, v in total_cells)
        for col, v in total_cells:
            if v == best_total:
                ws.cell(row=totals_row, column=col).fill = best_fill

    thin_side = Side(border_style="thin", color="000000")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    max_col = ws.max_column
    max_row = ws.max_row

    # --- Apply COA hyperlinks after all rows are written (keeps formats intact) ---
    for (r, c), url in cert_links.items():
        cell = ws.cell(row=r, column=c)
        cell.hyperlink = url

        # Make it look like hyperlink WITHOUT using cell.style = "Hyperlink"
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            underline="single",
            color="0000FF"
        )

    # ---------- A) Apply borders + enable wrapping on ALL cells ----------
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal=(cell.alignment.horizontal or "left"),
                vertical="center",
                wrap_text=True
            )

    # ---------- B) Compute & apply capped "autofit" column widths ----------
    col_max_len = [0] * (max_col + 1)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            text = "" if val is None else str(val)
            col_max_len[c] = max(col_max_len[c], len(text))

    DEFAULT_CAP = 25
    DIM_CAP = 12
    SUPPLIER_CAP = 18
    REMARKS_CAP = 30

    dim_start = 4
    dim_end = 3 + DIM_COUNT
    qty_col = dim_end + 1
    wt_col  = dim_end + 2

    def col_cap(c: int) -> int:
        if dim_start <= c <= dim_end:
            return DIM_CAP
        if c == qty_col or (has_weight and c == wt_col):
            return 12
        if c == 1:  # item name
            return 35
        if c == 2:  # brand/spec
            return 30
        if c == 3:  # category
            return 20
        if c >= SUPPLIER_START_COL:
            offset = (c - SUPPLIER_START_COL) % SUPPLIER_BLOCK_COLS
            if offset == OFF_REMARKS:
                return REMARKS_CAP
            return SUPPLIER_CAP
        return DEFAULT_CAP

    for c in range(1, max_col + 1):
        cap = col_cap(c)
        est = int(col_max_len[c] * 0.8) + 2
        width = max(8, min(cap, est))
        ws.column_dimensions[get_column_letter(c)].width = width

    # ---------- C) Now auto-fit ROW HEIGHT based on WRAPPED lines ----------
    def estimate_wrapped_lines(text: str, col_width: float) -> int:
        """
        Rough estimate: Excel column width ~ "characters".
        This isn't perfect, but works well enough for wrap+height.
        """
        if not text:
            return 1
        chars_per_line = max(1, int(col_width))  # rough
        total_lines = 0
        for part in text.split("\n"):
            total_lines += max(1, math.ceil(len(part) / chars_per_line))
        return total_lines

    for r in range(1, max_row + 1):
        max_lines = 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            text = "" if cell.value is None else str(cell.value)

            col_letter = get_column_letter(c)
            col_w = ws.column_dimensions[col_letter].width or 10

            lines = estimate_wrapped_lines(text, col_w)
            max_lines = max(max_lines, lines)

        ws.row_dimensions[r].height = max(15, max_lines * 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"comparison_task_{task_id}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/suppliers')
def suppliers():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get suppliers with their categories as comma-separated string
    suppliers_list = []
    all_suppliers = db.session.query(Supplier).filter_by(is_active=True).all()
    
    for supplier in all_suppliers:
        category_names = [cat.name for cat in supplier.categories]
        supplier_dict = {
            'id': supplier.id,
            'name': supplier.name,
            'contact_name': supplier.contact_name,
            'email': supplier.email,
            'contact_number': supplier.contact_number,
            'address': supplier.address,
            'products_services': supplier.products_services,
            'is_active': supplier.is_active,
            'categories': ', '.join(category_names) if category_names else ''
        }
        suppliers_list.append(supplier_dict)
    
    categories = db.session.query(Category).all()

    return render_template('supplier_list.html', 
                         suppliers=suppliers_list, 
                         categories=categories,
                         user_role=session['role'])

@app.route('/edit-supplier/<int:supplier_id>', methods=['GET', 'POST'])
def edit_supplier(supplier_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('suppliers'))
    
    if request.method == 'POST':
        name = request.form['name']
        contact_name = request.form['contact_name']
        email = request.form['email']
        contact_number = request.form['contact_number']
        address = request.form['address']
        products_services = request.form['products_services']
        selected_categories = request.form.getlist('categories')
        
        # Update supplier
        supplier = db.session.get(Supplier, supplier_id)
        if not supplier:
            flash('Supplier not found', 'error')
            return redirect(url_for('suppliers'))
        
        supplier.name = name
        supplier.contact_name = contact_name
        supplier.email = email
        supplier.contact_number = contact_number
        supplier.address = address
        supplier.products_services = products_services
        
        # Update categories - clear existing and add new ones
        supplier.categories.clear()
        for category_id in selected_categories:
            category = db.session.get(Category, category_id)
            if category:
                supplier.categories.append(category)
        
        try:
            db.session.commit()
            flash('Supplier updated successfully!', 'success')
            return redirect(url_for('suppliers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating supplier: {str(e)}', 'error')
    
    # GET request
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier:
        flash('Supplier not found', 'error')
        return redirect(url_for('suppliers'))
    
    categories = db.session.query(Category).all()
    supplier_category_ids = [cat.id for cat in supplier.categories]
    
    return render_template('edit_supplier.html', 
                         supplier=supplier, 
                         categories=categories,
                         supplier_category_ids=supplier_category_ids)

@app.route('/add-supplier', methods=['GET', 'POST'])
def add_supplier():
    if 'user_id' not in session:
        flash('Access denied', 'error')
        return redirect(url_for('suppliers'))
    
    if request.method == 'POST':
        name = request.form['name']
        contact_name = request.form['contact_name']
        email = request.form['email']
        contact_number = request.form['contact_number']
        address = request.form['address']
        products_services = request.form['products_services']
        selected_categories = request.form.getlist('categories')
        
        # Validate email
        if not validate_email(email):
            flash('Invalid email format', 'error')
            categories = db.session.query(Category).all()
            return render_template('edit_supplier.html', categories=categories)
        
        # Validate phone
        if contact_number and not validate_phone(contact_number):
            flash('Invalid phone number format', 'error')
            categories = db.session.query(Category).all()
            return render_template('edit_supplier.html', categories=categories)

        # Check for duplicate supplier (Name or Email)
        existing = db.session.query(Supplier).filter(
            (Supplier.name == name) | (Supplier.email == email)
        ).first()
        
        if existing:
            flash(f'Supplier already exists (Name: {existing.name}). Duplicate entry prevented.', 'error')
            categories = db.session.query(Category).all()
            return render_template('edit_supplier.html', categories=categories)
        
        # Create supplier
        supplier = Supplier(
            name=name,
            contact_name=contact_name,
            email=email,
            contact_number=contact_number,
            address=address,
            products_services=products_services,
            is_active=True
        )
        
        # Add categories
        for category_id in selected_categories:
            category = db.session.get(Category, category_id)
            if category:
                supplier.categories.append(category)
        
        try:
            db.session.add(supplier)
            db.session.commit()
            flash('Supplier added successfully!', 'success')
            return redirect(url_for('suppliers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding supplier: {str(e)}', 'error')
            categories = db.session.query(Category).all()
            return render_template('edit_supplier.html', categories=categories)
    
    # GET request
    categories = db.session.query(Category).all()
    return render_template('edit_supplier.html', categories=categories)

@app.route('/delete-task/<int:task_id>', methods=['POST'])
def delete_task(task_id):
    if 'user_id' not in session:
        flash('Access denied', 'error')
        return redirect(url_for('login'))

    try:
        # Just check it exists (no ownership check)
        task = db.session.get(Task, task_id)
        if not task:
            flash('Task not found', 'error')
            return redirect(url_for('task_list'))

        # Cascading delete should handle related records via relationships
        db.session.delete(task)
        db.session.commit()
        flash('Task deleted successfully!', 'success')
        return redirect(url_for('task_list'))

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Failed to delete task_id=%s: %s", task_id, str(e))
        flash(f'Failed to delete task: {str(e)}', 'error')
        return redirect(url_for('task_list'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('login'))

def _send_email_via_sendgrid(to_email: str, to_name: str, subject: str, html_body: str) -> bool:
    """
    Send via SendGrid. Returns True on success; False if not configured or failed.
    """
    if not (SENDGRID_API_KEY and SENDGRID_SENDER):
        return False

    try:
        # Often fixes Windows/proxy TLS issues
        os.environ["SSL_CERT_FILE"] = certifi.where()

        sg = SendGridAPIClient(SENDGRID_API_KEY)
        message = Mail(
            from_email=Email(SENDGRID_SENDER, "Procurement Department"),
            to_emails=To(to_email, to_name or ""),
            subject=subject,
            html_content=html_body,
        )
        app.logger.info("Sending email via SendGrid to=%s subject=%s", to_email, subject)
        resp = sg.send(message)
        return 200 <= resp.status_code < 300
    except Exception:
        app.logger.exception("SendGrid send failed")
        return False


def _send_email_via_smtp(to_email: str, subject: str, html_body: str) -> bool:
    """
    Send via SMTP. Returns True on success, False on failure.
    """
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html_body, 'html'))

        app.logger.info("Sending email via SMTP to=%s subject=%s", to_email, subject)

        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.starttls()
        server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
        server.send_message(msg)
        server.quit()
        return True
    except Exception:
        app.logger.exception("SMTP send failed")
        return False


def send_email_html(to_email: str, subject: str, html_body: str, to_name: str = "") -> bool:
    """
    Unified reusable email sender:
    - Try SendGrid first (if configured)
    - Fall back to SMTP
    """
    if _send_email_via_sendgrid(to_email, to_name, subject, html_body):
        return True
    return _send_email_via_smtp(to_email, subject, html_body)


# Email sending function
def send_procurement_email(supplier_email, supplier_name, pr_items, task_name, assigned_item_ids=None, custom_content=None, subject=None, supplier_contact=None, quote_form_link=None):
    try:
        # Filter items for this specific supplier
        if assigned_item_ids:
            assigned_ids = [int(item_id) for item_id in assigned_item_ids]
            supplier_items = [item for item in pr_items if item['id'] in assigned_ids]
        else:
            supplier_items = pr_items
        
        if not supplier_items:
            print(f"No items assigned to {supplier_name}, skipping email")
            return False
        
        subject = subject or f"Procurement Inquiry - {task_name}"
        
        if custom_content:
            body = custom_content.replace('{supplier_name}', supplier_name)
            # Replace contact person placeholder if present
            if supplier_contact:
                body = body.replace('{contact_person}', supplier_contact)
            else:
                body = body.replace('{contact_person}', '')

            if quote_form_link:
                quote_html = f'<p><strong>Quotation Form Link:</strong> <a href="{quote_form_link}">{quote_form_link}</a></p>'

                # If user put a placeholder in the template, replace it
                if '{quote_form_link}' in body:
                    body = body.replace('{quote_form_link}', quote_form_link)
                # Otherwise, try to insert nicely before </body>
                elif '</body>' in body:
                    body = body.replace('</body>', quote_html + '</body>')
                else:
                    body += quote_html
        else:
            items_html = """
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; border: 1px solid #ddd; font-family: Arial, sans-serif;">
                <thead style="background-color: #f2f2f2;">
                    <tr>
                        <th style="text-align: center; width: 5%;">No.</th>
                        <th style="text-align: left;">Description</th>
                        <th style="text-align: left;">Dimensions</th>
                        <th style="text-align: left;">Brand / Specification</th>
                        <th style="text-align: center; width: 10%;">Qty</th>
                        <th style="text-align: left;">Remark</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for idx, item in enumerate(pr_items, 1):
                item = dict(item) if item is not None else {}
                category = (item.get('item_category') or '').strip()

                # Build Specification (Dimensions)
                spec = ''
                if category in ['Steel Plates', 'Stainless Steel']:
                    w = item.get('width') or ''
                    l = item.get('length') or ''
                    thk = item.get('thickness') or ''
                    if w or l or thk:
                        spec = f"{w} mm (W) x {l} mm (L) x {thk} mm (Thk)"
                elif category == 'Angle Bar':
                    a = item.get('dim_a') or ''
                    b = item.get('dim_b') or ''
                    l = item.get('length') or ''
                    thk = item.get('thickness') or ''
                    if a or b or l or thk:
                        spec = f"{a} mm (A) x {b} mm (B) x {l} mm (L) x {thk} mm (Thk)"
                elif category in ['Rebar', 'Bolts, Fasteners']:
                    d = item.get('diameter') or ''
                    l = item.get('length') or ''
                    if d or l:
                        spec = f"{d} mm (D) x {l} mm (L)"
                else:
                    uom = item.get('uom') or ''
                    uom_qty = item.get('uom_qty') or ''
                    if uom_qty and uom:
                        spec = f"{uom_qty} {uom}"
                    elif uom_qty:
                        spec = f"Qty: {uom_qty}"
                    elif uom:
                        spec = f"UOM: {uom}"
                
                # Fallback to direct spec field if dynamic dim is empty, or append?
                # User image shows "Specification" column often having sizes. try to use 'specification' field if dims are empty?
                # Actually in pr_items we have 'specification' column.
                original_spec = item.get('specification') or ''
                if spec and original_spec:
                    final_spec = f"{spec}<br><small>{original_spec}</small>"
                elif spec:
                    final_spec = spec
                else:
                    final_spec = original_spec or 'N/A'

                items_html += f"""
                    <tr>
                        <td style="text-align: center;">{idx}</td>
                        <td>{item['item_name']}</td>
                        <td>{final_spec}</td>
                        <td>{item.get('brand') or ''}</td>
                        <td style="text-align: center;">{item['quantity']}</td>
                        <td></td> <!-- Empty Remark for now, as in image -->
                    </tr>
                """


            items_html += "</tbody></table>"
            
            body = f"""
            <html>
            <body>
                <h2>Procurement Inquiry</h2>
                <p>Dear {(supplier_name.strip() + ' from ' + supplier_contact.strip()) if (supplier_contact and supplier_contact.strip()) else (supplier_name.strip() if supplier_name else '')},</p>
                
                <p>We are inquiring about the following items for procurement:</p>
                
                {items_html}
                
                <p>Please provide us with your quotation including:</p>
                <ul>
                    <li>Payment terms</li>
                    <li>Unit Price (RM))</li>
                    <li>Delivery Lead Timeline</li>
                    <li>Stock Availability</li>
                    <li>Warranty (If Applicable)</li>
                    <li>Mill Certificate / Certificate of Analysis (COA)</li>
                </ul>

                <p>Please fill in the quotation in the link below:</p>
                <p>Supplier form: </p>
                <p>↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓</p>
                {quote_form_link}
                <p>↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑</p>
                
                <p>We look forward to your prompt response.</p>
                
                <p>Best regards,<br>
                Procurement Department</p>
            </body>
            </html>
            """

        ok = send_email_html(
            to_email=supplier_email,
            to_name=supplier_name or "",
            subject=subject,
            html_body=body
        )
        return ok

    except Exception:
        app.logger.exception("Email sending failed")
        return False

@app.route('/task/<int:task_id>/suppliers', methods=['GET', 'POST'])
def supplier_selection(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verify task ownership
    task = db.session.get(Task, task_id)
    if not task:
        flash('Task not found', 'error')
        return redirect(url_for('task_list'))
    
    if request.method == 'POST':
        selected_suppliers = request.form.getlist('suppliers')
        
        # Clear existing selections
        db.session.query(TaskSupplier).filter_by(task_id=task_id).delete()
        
        # Add supplier selections
        for supplier_id in selected_suppliers:
            task_supplier = TaskSupplier(
                task_id=task_id,
                supplier_id=supplier_id,
                is_selected=True,
                assigned_items=None
            )
            db.session.add(task_supplier)
        
        # Update task status
        task.status = 'select_suppliers'
        
        try:
            db.session.commit()
            flash('Suppliers and item assignments saved successfully!', 'success')
            return redirect(url_for('email_preview', task_id=task_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving suppliers: {str(e)}', 'error')
    
    # Get PR items
    pr_items = db.session.query(PRItem).filter_by(task_id=task_id).all()
    
    # Get unique categories from PR items
    categories = list(set([item.item_category for item in pr_items]))
    
    # Get suppliers that match ANY of the categories
    suppliers = []
    if categories:
        suppliers = db.session.query(Supplier).join(
            Supplier.categories
        ).filter(
            Category.name.in_(categories),
            Supplier.is_active == True
        ).distinct().all()
    
    # Get already selected suppliers
    selected_data = db.session.query(TaskSupplier).filter_by(
        task_id=task_id,
        is_selected=True
    ).all()
    
    selected_supplier_ids = [str(data.supplier_id) for data in selected_data]
    
    return render_template('supplier_selection.html', 
                         task=task, 
                         suppliers=suppliers, 
                         pr_items=pr_items,
                         selected_supplier_ids=selected_supplier_ids,
                         categories=categories)

@app.route('/delete-supplier/<int:supplier_id>')
def delete_supplier(supplier_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('suppliers'))
    
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier:
        flash('Supplier not found', 'error')
        return redirect(url_for('suppliers'))
    
    # Soft delete by setting is_active to 0
    supplier.is_active = False
    
    try:
        db.session.commit()
        flash('Supplier deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting supplier: {str(e)}', 'error')
    
    return redirect(url_for('suppliers'))

# Add route for categories management (view for all, edit for admin)
@app.route('/categories')
def categories():
    if 'user_id' not in session:
        flash('Please login first', 'error')
        return redirect(url_for('login'))
    
    can_edit = session.get('role') == 'admin'
    
    categories_list = db.session.query(Category).order_by(Category.name).all()
    
    return render_template('categories.html', categories=categories_list, can_edit=can_edit)

@app.route('/add-category', methods=['POST'])
def add_category():
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403
    
    name = request.form.get('name')
    if not name:
        flash('Category name is required', 'error')
        return redirect(url_for('categories'))
    
    # Check if category exists
    existing = db.session.query(Category).filter_by(name=name).first()
    if existing:
        flash('Category already exists', 'error')
        return redirect(url_for('categories'))
    
    try:
        category = Category(name=name)
        db.session.add(category)
        db.session.commit()
        flash('Category added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding category: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

@app.route('/delete-category/<int:category_id>')
def delete_category(category_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('categories'))
    
    category = db.session.get(Category, category_id)
    if not category:
        flash('Category not found', 'error')
        return redirect(url_for('categories'))
    
    # Check if category is used by any suppliers
    suppliers_count = db.session.query(Supplier).filter(
        Supplier.categories.any(id=category_id)
    ).count()
    
    if suppliers_count > 0:
        flash('Cannot delete category: it is being used by suppliers', 'error')
    else:
        try:
            db.session.delete(category)
            db.session.commit()
            flash('Category deleted successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting category: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

@app.route('/category/<int:category_id>/items')
def category_items(category_id):
    if 'user_id' not in session:
        flash('Please login first', 'error')
        return redirect(url_for('login'))
    
    can_edit = session.get('role') == 'admin'
    
    category = db.session.get(Category, category_id)
    if not category:
        flash('Category not found', 'error')
        return redirect(url_for('categories'))
        
    items = db.session.query(CategoryItem).filter_by(category_id=category_id).order_by(CategoryItem.name).all()
    
    return render_template('category_items.html', category=category, items=items, can_edit=can_edit)

@app.route('/category/<int:category_id>/add-item', methods=['POST'])
def add_category_item(category_id):
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    name = request.form.get('name')
    if name:
        try:
            item = CategoryItem(category_id=category_id, name=name)
            db.session.add(item)
            db.session.commit()
            flash('Item added successfully', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding item: {str(e)}', 'error')
    
    return redirect(url_for('category_items', category_id=category_id))

@app.route('/delete-category-item/<int:item_id>')
def delete_category_item(item_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    item = db.session.get(CategoryItem, item_id)
    if item:
        category_id = item.category_id
        try:
            db.session.delete(item)
            db.session.commit()
            flash('Item deleted successfully', 'success')
            return redirect(url_for('category_items', category_id=category_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting item: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

def parse_reply_fields(body_text):
    """
    Try to extract:
      - unit price
      - total price
      - delivery timeline / lead time
      - warranty info
      - payment terms
    from a reasonably structured reply.

    Expected style (case-insensitive, flexible about spaces):
        Unit price: ...
        Total price: ...
        Delivery timeline: ...
        Warranty information: ...
        Payment terms: ...
    """
    # Normalize line endings a bit
    text = body_text.replace('\r\n', '\n').replace('\r', '\n')

    def extract(pattern):
        """
        Return the first capture group for regex `pattern` in text,
        or None if not found.
        """
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            return None
        return m.group(1).strip()

    unit_price = extract(r"unit\s+price\s*[:\-]\s*([^\n\r]+)")
    stock_availability = extract(r"total\s+price\s*[:\-]\s*([^\n\r]+)")
    lead_time = extract(r"(?:delivery\s+timeline|lead\s+time)\s*[:\-]\s*([^\n\r]+)")
    warranty = extract(r"warranty\s+information\s*[:\-]\s*([^\n\r]+)")
    payment_terms = extract(r"payment\s+terms\s*[:\-]\s*([^\n\r]+)")

    return {
        "unit_price": unit_price,
        "stock_availability": stock_availability,
        "lead_time": lead_time,
        "warranty": warranty,
        "payment_terms": payment_terms,
    }

def get_email_body(msg):
    """Extract text/plain part (or fallback) from an email.message.Message."""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition") or "")
            if ctype == "text/plain" and "attachment" not in disp:
                charset = part.get_content_charset() or "utf-8"
                return part.get_payload(decode=True).decode(charset, errors="ignore")
        # Fallback: no text/plain found
        for part in msg.walk():
            if part.get_content_type().startswith("text/"):
                charset = part.get_content_charset() or "utf-8"
                return part.get_payload(decode=True).decode(charset, errors="ignore")
        return ""
    else:
        charset = msg.get_content_charset() or "utf-8"
        try:
            return msg.get_payload(decode=True).decode(charset, errors="ignore")
        except Exception:
            return str(msg.get_payload())

# ==================================== DEBUG ====================================
# Add this route to your app.py temporarily
@app.route('/debug-db')
def debug_db():
    if not ENABLE_DEBUG_ROUTES:
        abort(404)
    if 'user_id' not in session or session.get('role') != 'admin':
        abort(403)

    conn = get_db_connection()
    
    # Check all tables
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    result = "<h1>Database Debug Info</h1>"
    result += f"<p>Tables found: {[table[0] for table in tables]}</p>"
    
    # Check each table's structure and content
    for table in tables:
        table_name = table[0]
        result += f"<h2>Table: {table_name}</h2>"
        
        # Get table structure
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        result += f"<p>Columns: {[col[1] for col in columns]}</p>"
        
        # Get table content
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        result += f"<p>Row count: {len(rows)}</p>"
        if rows:
            result += "<ul>"
            for row in rows:
                result += f"<li>{row}</li>"
            result += "</ul>"
    
    conn.close()
    return result

@app.route('/debug-info')
def debug_info():
    if not ENABLE_DEBUG_ROUTES:
        abort(404)
    if 'user_id' not in session or session.get('role') != 'admin':
        abort(403)

    import os
    import sqlite3
    
    current_dir = os.getcwd()
    db_path = os.path.join(current_dir, 'database', 'procure_flow.db')
    
    info = f"""
    <h1>Debug Information</h1>
    <p>Current directory: {current_dir}</p>
    <p>Database path: {db_path}</p>
    <p>Database exists: {os.path.exists(db_path)}</p>
    """
    
    if os.path.exists(db_path):
        info += f"<p>Database size: {os.path.getsize(db_path)} bytes</p>"
        
        # Check tables and row counts
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        info += "<h2>Tables and Row Counts:</h2>"
        for table in tables:
            table_name = table[0]
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            info += f"<p>{table_name}: {count} rows</p>"
        
        conn.close()
    
    return info

@app.route('/db-location')
def db_location():
    if not ENABLE_DEBUG_ROUTES:
        abort(404)
    if 'user_id' not in session or session.get('role') != 'admin':
        abort(403)

    import os
    import sqlite3
    
    current_dir = os.getcwd()
    db_path = os.path.abspath(os.path.join('database', 'procure_flow.db'))
    
    info = f"""
    <h1>Database Location</h1>
    <p><strong>Current working directory:</strong> {current_dir}</p>
    <p><strong>Database absolute path:</strong> {db_path}</p>
    <p><strong>Database exists:</strong> {os.path.exists(db_path)}</p>
    """
    
    if os.path.exists(db_path):
        # Show file properties
        import datetime
        stat = os.stat(db_path)
        modified_time = datetime.datetime.fromtimestamp(stat.st_mtime)
        
        info += f"""
        <p><strong>Database size:</strong> {stat.st_size} bytes</p>
        <p><strong>Last modified:</strong> {modified_time}</p>
        <p><strong>Full path to open in DB Browser:</strong></p>
        <code>{db_path}</code>
        """
    
    return info
# ==================================== END OF DEBUG ====================================

@app.errorhandler(404)
def not_found(e):
    return render_template("errors/404.html"), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template("errors/403.html"), 403

@app.errorhandler(400)
def bad_request(e):
    return render_template("errors/400.html"), 400

@app.errorhandler(500)
def server_error(e):
    # Log the real error internally
    app.logger.exception("500 Server Error")
    return render_template("errors/500.html"), 500

@app.errorhandler(Exception)
def unhandled_exception(e):
    if isinstance(e, HTTPException):
        return e
    app.logger.exception("Unhandled exception")
    return render_template("errors/500.html"), 500

if __name__ == '__main__':
    # for Render: always use 0.0.0.0
    debug_mode = os.getenv("FLASK_DEBUG", "0") == "1" and os.getenv("RENDER", "0") != "1"
    host = '0.0.0.0'  # Force 0.0.0.0 for Render
    port = int(os.getenv("PORT", "5000"))
    
    app.run(host=host, port=port, debug=debug_mode)
